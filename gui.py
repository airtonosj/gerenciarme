import customtkinter as ctk
from tkcalendar import Calendar
from tkinter import messagebox, ttk, filedialog
import threading
import os
import pandas as pd
import numpy as np
from bot_logic import BotController
import matplotlib
matplotlib.use("TkAgg")
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from matplotlib.figure import Figure
import matplotlib.gridspec as gridspec
from datetime import datetime, timedelta
from db_manager import importar_lancamentos_diarios, buscar_lancamentos_por_periodo, buscar_equipamentos_sem_lancamento
from tkcalendar import Calendar, DateEntry # <-- Adicione DateEntry aqui


# ─── Palette ──────────────────────────────────────────────────────────────────
DARK = {
    "bg":        "#0f1117",
    "surface":   "#1a1d27",
    "card":      "#21253a",
    "border":    "#2d3150",
    "accent":    "#4f8ef7",
    "accent2":   "#7b5ea7",
    "success":   "#3ecf8e",
    "warn":      "#f5a623",
    "danger":    "#e84545",
    "text":      "#e8eaf6",
    "subtext":   "#8892b0",
    "mpl_bg":    "#1a1d27",
    "mpl_grid":  "#2d3150",
    "mpl_text":  "#e8eaf6",
}
LIGHT = {
    "bg":        "#f0f2f8",
    "surface":   "#ffffff",
    "card":      "#f7f8fc",
    "border":    "#d1d5e8",
    "accent":    "#3a6fd8",
    "accent2":   "#6b47a8",
    "success":   "#1ea972",
    "warn":      "#d4861a",
    "danger":    "#c93232",
    "text":      "#1a1d2e",
    "subtext":   "#5c6482",
    "mpl_bg":    "#ffffff",
    "mpl_grid":  "#e0e4f0",
    "mpl_text":  "#1a1d2e",
}

CHART_COLORS = ["#4f8ef7", "#3ecf8e", "#f5a623", "#e84545", "#7b5ea7",
                "#00c8d7", "#ff6b6b", "#ffd93d", "#6bcb77", "#c77dff"]


def hex_to_rgb(h):
    h = h.lstrip("#")
    return tuple(int(h[i:i+2], 16)/255 for i in (0, 2, 4))


def _tint(hex_color, alpha=0.18, bg_key="card"):
    """Blend hex_color onto the theme background at `alpha` opacity → solid hex."""
    def parse(h):
        h = h.lstrip("#")
        return tuple(int(h[i:i+2], 16) for i in (0, 2, 4))
    try:
        fr, fg, fb = parse(hex_color)
        br, bg, bb = parse(ThemeManager.p(bg_key))
        r = int(fr * alpha + br * (1 - alpha))
        g = int(fg * alpha + bg * (1 - alpha))
        b = int(fb * alpha + bb * (1 - alpha))
        return f"#{r:02x}{g:02x}{b:02x}"
    except Exception:
        return hex_color


class ThemeManager:
    _mode = "dark"
    _palette = DARK
    _listeners = []

    @classmethod
    def toggle(cls):
        cls._mode = "light" if cls._mode == "dark" else "dark"
        cls._palette = LIGHT if cls._mode == "light" else DARK
        ctk.set_appearance_mode("Light" if cls._mode == "light" else "Dark")
        for cb in cls._listeners:
            cb()

    @classmethod
    def p(cls, key):
        return cls._palette[key]

    @classmethod
    def register(cls, cb):
        cls._listeners.append(cb)

    @classmethod
    def is_dark(cls):
        return cls._mode == "dark"


# ─── Re-usable widgets ────────────────────────────────────────────────────────

class ModernCard(ctk.CTkFrame):
    def __init__(self, master, **kwargs):
        kwargs.setdefault("corner_radius", 14)
        kwargs.setdefault("border_width", 1)
        super().__init__(master, **kwargs)
        self._sync()
        ThemeManager.register(self._sync)

    def _sync(self):
        self.configure(fg_color=ThemeManager.p("card"),
                       border_color=ThemeManager.p("border"))


class GlowButton(ctk.CTkButton):
    def __init__(self, master, color_key="accent", **kwargs):
        self._ck = color_key
        kwargs.setdefault("corner_radius", 10)
        kwargs.setdefault("height", 42)
        kwargs.setdefault("font", ctk.CTkFont("Segoe UI", 13, weight="bold"))
        super().__init__(master, **kwargs)
        self._sync()
        ThemeManager.register(self._sync)

    def _sync(self):
        c = ThemeManager.p(self._ck)
        self.configure(fg_color=c, hover_color=self._darken(c))

    @staticmethod
    def _darken(h, factor=0.8):
        r, g, b = hex_to_rgb(h)
        return "#{:02x}{:02x}{:02x}".format(int(r*factor*255), int(g*factor*255), int(b*factor*255))


class StatusBadge(ctk.CTkLabel):
    ICONS = {"ok": "✓", "fail": "✗", "wait": "◌", "info": "●"}

    def __init__(self, master, text, kind="info", **kwargs):
        icon = self.ICONS.get(kind, "●")
        colors = {"ok": "success", "fail": "danger", "wait": "warn", "info": "accent"}
        color = ThemeManager.p(colors.get(kind, "accent"))
        super().__init__(master,
                         text=f" {icon}  {text} ",
                         corner_radius=8,
                         fg_color=_tint(color, alpha=0.22),
                         text_color=color,
                         font=ctk.CTkFont("Segoe UI", 11),
                         **kwargs)


# ─── Tab 1 – Bot Controller ────────────────────────────────────────────────────

class BotTab(ctk.CTkFrame):
    def __init__(self, master):
        super().__init__(master, fg_color="transparent")
        self.start_date = None
        self.end_date = None
        self._build()
        ThemeManager.register(self._sync_theme)

    def _build(self):
        self.grid_columnconfigure(0, weight=0, minsize=340)
        self.grid_columnconfigure(1, weight=1)
        self.grid_rowconfigure(0, weight=1)

        # ── Left panel ──
        left = ModernCard(self)
        left.grid(row=0, column=0, padx=(0, 10), pady=0, sticky="nsew")
        left.grid_propagate(False)

        # Title
        ctk.CTkLabel(left, text="🤖  Controles do Bot",
                     font=ctk.CTkFont("Segoe UI", 17, weight="bold"),
                     anchor="w").pack(pady=(20, 5), padx=20, fill="x")

        ctk.CTkLabel(left, text="Selecione o período e inicie a coleta automática.",
                     font=ctk.CTkFont("Segoe UI", 11),
                     text_color=ThemeManager.p("subtext"),
                     anchor="w").pack(padx=20, fill="x")

        sep = ctk.CTkFrame(left, height=1, fg_color=ThemeManager.p("border"))
        sep.pack(fill="x", padx=20, pady=12)

        # Browser checkbox
        self.show_browser_var = ctk.BooleanVar(value=True)
        chk = ctk.CTkCheckBox(left, text="Exibir navegador durante execução",
                              variable=self.show_browser_var,
                              font=ctk.CTkFont("Segoe UI", 12))
        chk.pack(padx=20, anchor="w", pady=(0, 14))

        # Calendar
        style = ttk.Style()
        style.theme_use("clam")
        self._cal_style_key = "Bot.TCalendar"
        style.configure(self._cal_style_key,
                        background="#1a1d27", foreground="#e8eaf6",
                        fieldbackground="#1a1d27",
                        selectbackground="#4f8ef7",
                        weekendforeground="#7b5ea7",
                        headersbackground="#21253a")

        self.cal = Calendar(left, selectmode="day", locale="pt_BR",
                            style=self._cal_style_key,
                            font="Arial 10",
                            showweeknumbers=False)
        self.cal.pack(padx=20, fill="x")
        self.cal.bind("<<CalendarSelected>>", self._on_date_select)

        # Date range labels
        dates_row = ctk.CTkFrame(left, fg_color="transparent")
        dates_row.pack(pady=8, padx=20, fill="x")
        dates_row.grid_columnconfigure((0, 1), weight=1)

        self.lbl_de = ctk.CTkLabel(dates_row, text="De: --/--/----",
                                   font=ctk.CTkFont("Segoe UI", 12, weight="bold"),
                                   text_color=ThemeManager.p("accent"))
        self.lbl_de.grid(row=0, column=0, sticky="w")
        self.lbl_para = ctk.CTkLabel(dates_row, text="Para: --/--/----",
                                     font=ctk.CTkFont("Segoe UI", 12, weight="bold"),
                                     text_color=ThemeManager.p("accent2"))
        self.lbl_para.grid(row=0, column=1, sticky="e")

        sep2 = ctk.CTkFrame(left, height=1, fg_color=ThemeManager.p("border"))
        sep2.pack(fill="x", padx=20, pady=10)

        self.start_btn = GlowButton(left, text="▶  Iniciar Bot",
                                    command=self._start_bot_thread,
                                    color_key="accent")
        self.start_btn.pack(padx=20, fill="x", pady=(0, 6))

        ctk.CTkLabel(left, text="📁  Relatórios salvos na pasta Downloads",
                     font=ctk.CTkFont("Segoe UI", 10),
                     text_color=ThemeManager.p("subtext"),
                     wraplength=280, anchor="w").pack(padx=20, pady=(6, 20), fill="x")

        # ── Right panel ──
        right = ModernCard(self)
        right.grid(row=0, column=1, padx=(10, 0), pady=0, sticky="nsew")
        right.grid_rowconfigure(1, weight=1)

        hdr = ctk.CTkFrame(right, fg_color="transparent")
        hdr.pack(fill="x", padx=20, pady=(18, 8))
        ctk.CTkLabel(hdr, text="📋  Status da Execução",
                     font=ctk.CTkFont("Segoe UI", 15, weight="bold")).pack(side="left")

        self.status_count = ctk.CTkLabel(hdr, text="0 eventos",
                                          font=ctk.CTkFont("Segoe UI", 11),
                                          text_color=ThemeManager.p("subtext"))
        self.status_count.pack(side="right")

        # Treeview
        style.configure("Bot.Treeview",
                        background="#1a1d27", foreground="#e8eaf6",
                        rowheight=28, fieldbackground="#1a1d27",
                        borderwidth=0, font=("Segoe UI", 11))
        style.map("Bot.Treeview", background=[("selected", "#4f8ef7")])
        style.configure("Bot.Treeview.Heading",
                        background="#21253a", foreground="#8892b0",
                        relief="flat", font=("Segoe UI", 11, "bold"))

        tree_frame = ctk.CTkFrame(right, fg_color="transparent")
        tree_frame.pack(fill="both", expand=True, padx=20, pady=(0, 20))

        self.log_tree = ttk.Treeview(tree_frame, style="Bot.Treeview",
                                     columns=("ts", "msg"), show="headings")
        self.log_tree.heading("ts", text="Horário")
        self.log_tree.heading("msg", text="Mensagem")
        self.log_tree.column("ts", width=90, anchor="center")
        self.log_tree.column("msg", anchor="w")

        sb = ttk.Scrollbar(tree_frame, orient="vertical", command=self.log_tree.yview)
        self.log_tree.configure(yscrollcommand=sb.set)
        self.log_tree.pack(side="left", fill="both", expand=True)
        sb.pack(side="right", fill="y")

    def _sync_theme(self):
        pass  # calendar and treeview styles are mostly fixed; CTk handles the rest

    def _on_date_select(self, _):
        sel = self.cal.selection_get()
        if self.start_date is None or self.end_date is not None:
            self.start_date = sel
            self.end_date = None
            self.lbl_de.configure(text=f"De: {sel.strftime('%d/%m/%Y')}")
            self.lbl_para.configure(text="Para: --/--/----")
        else:
            if sel < self.start_date:
                self.end_date, self.start_date = self.start_date, sel
            else:
                self.end_date = sel
            self.lbl_de.configure(text=f"De: {self.start_date.strftime('%d/%m/%Y')}")
            self.lbl_para.configure(text=f"Para: {self.end_date.strftime('%d/%m/%Y')}")

    def _start_bot_thread(self):
        if not self.start_date or not self.end_date:
            messagebox.showerror("Erro", "Selecione uma data de início E uma data de fim.")
            return
        self.start_btn.configure(state="disabled", text="⏳  Executando…")
        self.log_tree.delete(*self.log_tree.get_children())
        self._event_count = 0

        params = {
            "start_date": self.start_date.strftime("%d/%m/%Y"),
            "end_date":   self.end_date.strftime("%d/%m/%Y"),
            "headless":   not self.show_browser_var.get(),
            "gui_instance": self,
        }
        threading.Thread(target=self._run_bot, args=(params,), daemon=True).start()

    def _run_bot(self, params):
        try:
            bot = BotController(**params)
            for status in bot.run():
                self.after(0, self._append_log, status)
            
            # ── GATILHO DE IMPORTAÇÃO AUTOMÁTICA PARA O MYSQL ──
            try:
                # O nome do arquivo gerado pelo bot_logic.py ainda é cravado com "ontem" (intocável)
                yesterday_str = (datetime.now() - timedelta(1)).strftime('%d-%m-%Y')
                download_path = os.path.join(os.path.expanduser('~'), 'Downloads')
                final_filename = f"Equipamentos Gerenciarme {yesterday_str}.csv"
                final_filepath = os.path.join(download_path, final_filename)
                
                # A DATA DE REFERÊNCIA PARA O BANCO agora é a data selecionada no calendário do Bot!
                # Pega o 'start_date' que vem no formato DD/MM/YYYY e converte para o YYYY-MM-DD do MySQL
                data_selecionada = datetime.strptime(params['start_date'], "%d/%m/%Y")
                db_date_ref = data_selecionada.strftime('%Y-%m-%d')
                
                if os.path.exists(final_filepath):
                    self.after(0, self._append_log, f"⚡ Banco de Dados: Lendo arquivo gerado...")
                    df_auto = pd.read_csv(final_filepath, sep=';', encoding='utf-8-sig', on_bad_lines='skip')
                    if df_auto.shape[1] == 1:
                        df_auto = pd.read_csv(final_filepath, sep=',', encoding='utf-8', on_bad_lines='skip')
                    
                    self.after(0, self._append_log, f"⚡ Banco de Dados: Gravando registros para a data {params['start_date']}...")
                    
                    if importar_lancamentos_diarios(df_auto, db_date_ref):
                        self.after(0, self._append_log, f"✓ Banco de Dados: Dados de {params['start_date']} salvos com sucesso!")
                    else:
                        self.after(0, self._append_log, "✗ Banco de Dados: Falha na gravação automática.")
            except Exception as db_err:
                self.after(0, self._append_log, f"⚠️ Erro ao salvar no banco: {db_err}")
            # ──────────────────────────────────────────────────

            self.after(0, self._finish)
        except Exception as e:
            self.after(0, messagebox.showerror, "Erro Crítico", str(e))
            self.after(0, self._finish)

    def _append_log(self, msg):
        import datetime
        ts = datetime.datetime.now().strftime("%H:%M:%S")
        tag = "ok" if any(k in msg.lower() for k in ["conclu", "sucesso", "download"]) else \
              "fail" if any(k in msg.lower() for k in ["falh", "erro", "timeout"]) else "info"
        self.log_tree.insert("", "end", values=(ts, msg), tags=(tag,))
        self.log_tree.tag_configure("ok",   foreground="#3ecf8e")
        self.log_tree.tag_configure("fail", foreground="#e84545")
        self.log_tree.tag_configure("info", foreground="#e8eaf6")
        self.log_tree.yview_moveto(1)
        self._event_count = self._event_count + 1 if hasattr(self, "_event_count") else 1
        self.status_count.configure(text=f"{self._event_count} eventos")

    def _finish(self):
        self.start_btn.configure(state="normal", text="▶  Iniciar Bot")


# ─── Tab 2 – Rateio ────────────────────────────────────────────────────────────

COL_MAP = {
    "Equipamento":  "Equipamento",
    "Hrs Manu":     "Hrs Manutenção",
    "Hrs Trab":     "Hrs Trabalhadas",
    "Hrs Disp":     "Hrs Disponíveis",
    "Hrs Extras":   "Hrs Extras",
    "$ Total Trab": "$ Total Trabalhadas",
    "$ Total Disp": "$ Total Disponíveis",
    "% Disp":       "% Disponível",
    "% Util":       "% Utilização",
    "CR":           "CR",
}
HR_COLS  = ["Hrs Manu", "Hrs Trab", "Hrs Disp"]
COST_COLS = ["$ Total Trab", "$ Total Disp"]
PCT_COLS  = ["% Disp", "% Util"]
NUM_COLS  = HR_COLS + COST_COLS + PCT_COLS


class RateioTab(ctk.CTkFrame):
    def __init__(self, master):
        super().__init__(master, fg_color="transparent")
        self.df_raw = None
        self.df_group = None
        self.df_compare = None
        self.df_compare_group = None
        self.df_base = None          # planilha base/gabarito para comparação de lançamentos
        self._charts = []
        self._build()
        ThemeManager.register(self._sync_theme)

    # ── Build ──────────────────────────────────────────────────────────────────
    def _build(self):
        self.grid_rowconfigure(1, weight=1)
        self.grid_columnconfigure(0, weight=1)

        # ── Top toolbar ──
        toolbar = ModernCard(self)
        toolbar.grid(row=0, column=0, sticky="ew", pady=(0, 10))

        inner = ctk.CTkFrame(toolbar, fg_color="transparent")
        inner.pack(fill="x", padx=20, pady=14)

        ctk.CTkLabel(inner, text="📊  Rateio & Análise",
                     font=ctk.CTkFont("Segoe UI", 17, weight="bold")).pack(side="left")

        self.lbl_file = ctk.CTkLabel(inner, text="Nenhum arquivo carregado",
                                     font=ctk.CTkFont("Segoe UI", 11),
                                     text_color=ThemeManager.p("subtext"))
        self.lbl_file.pack(side="left", padx=20)

        # Botões e Inputs da barra superior
        self.export_btn = GlowButton(inner, text="📥  Exportar Aba",
                   command=self._export_current_tab, color_key="success",
                   width=135, height=36)
        self.export_btn.pack(side="right", padx=(0, 8))
        self.export_btn.configure(state="disabled")

        GlowButton(inner, text="📂  Carregar .xlsx",
                   command=self._load_file, color_key="accent",
                   width=140, height=36).pack(side="right", padx=(0, 15))

        self.lbl_base = ctk.CTkLabel(inner, text="Base: nenhuma",
                                     font=ctk.CTkFont("Segoe UI", 10),
                                     text_color=ThemeManager.p("subtext"))
        self.lbl_base.pack(side="right", padx=(0, 4))

        GlowButton(inner, text="🗂️  Carregar Base",
                   command=self._load_base_file, color_key="warn",
                   width=140, height=36).pack(side="right", padx=(0, 6))

        # ── NOVOS CONTROLES DO BANCO DE DADOS ──
        self.btn_load_db = GlowButton(inner, text="🔍  Buscar",
                                      command=self._load_from_database, color_key="accent2",
                                      width=100, height=36)
        self.btn_load_db.pack(side="right", padx=(0, 5))

        # Calendário "Para" (Fim)
        self.cal_para = DateEntry(inner, width=10, background='#4f8ef7', foreground='white', 
                                  borderwidth=0, date_pattern='dd/mm/yyyy', font=('Segoe UI', 11))
        self.cal_para.pack(side="right", padx=(0, 5))
        ctk.CTkLabel(inner, text="Para:", font=ctk.CTkFont("Segoe UI", 11)).pack(side="right", padx=(0, 5))

        # Calendário "De" (Início)
        self.cal_de = DateEntry(inner, width=10, background='#4f8ef7', foreground='white', 
                                borderwidth=0, date_pattern='dd/mm/yyyy', font=('Segoe UI', 11))
        self.cal_de.pack(side="right", padx=(0, 5))
        ctk.CTkLabel(inner, text="De:", font=ctk.CTkFont("Segoe UI", 11)).pack(side="right", padx=(0, 10))
        
        ctk.CTkLabel(inner, text="Período do Banco:", font=ctk.CTkFont("Segoe UI", 11, weight="bold")).pack(side="right", padx=(0, 5))

        # ── Sub-tab notebook (A PARTE QUE HAVIA SUMIDO) ──
        self.nb = ttk.Notebook(self)
        self.nb.grid(row=1, column=0, sticky="nsew")
        self._style_notebook()

        self.tab_geral    = ctk.CTkFrame(self.nb, fg_color=ThemeManager.p("surface"))
        self.tab_horas    = ctk.CTkFrame(self.nb, fg_color=ThemeManager.p("surface"))
        self.tab_custos   = ctk.CTkFrame(self.nb, fg_color=ThemeManager.p("surface"))
        self.tab_pct      = ctk.CTkFrame(self.nb, fg_color=ThemeManager.p("surface"))
        self.tab_equip    = ctk.CTkFrame(self.nb, fg_color=ThemeManager.p("surface"))
        self.tab_lanc     = ctk.CTkFrame(self.nb, fg_color=ThemeManager.p("surface"))
        self.tab_resumo   = ctk.CTkFrame(self.nb, fg_color=ThemeManager.p("surface"))

        self.nb.add(self.tab_geral,  text="  📋 Geral  ")
        self.nb.add(self.tab_horas,  text="  ⏱️ Horas  ")
        self.nb.add(self.tab_custos, text="  💰 Custos  ")
        self.nb.add(self.tab_pct,    text="  📈 Percentuais  ")
        self.nb.add(self.tab_equip,  text="  🏭 Equipamentos  ")
        self.nb.add(self.tab_lanc,   text="  🔔 Lançamentos  ")
        self.nb.add(self.tab_resumo, text="  🗂️ Resumo  ")

        self._placeholder("Carregue um arquivo .xlsx ou busque no banco para visualizar os dados agrupados por CR.")

    def _style_notebook(self):
        s = ttk.Style()
        bg  = ThemeManager.p("surface")
        acc = ThemeManager.p("accent")
        txt = ThemeManager.p("text")
        s.configure("Rateio.TNotebook", background=bg, borderwidth=0)
        s.configure("Rateio.TNotebook.Tab",
                    background=ThemeManager.p("card"),
                    foreground=ThemeManager.p("subtext"),
                    padding=[14, 8],
                    font=("Segoe UI", 11))
        s.map("Rateio.TNotebook.Tab",
              background=[("selected", bg)],
              foreground=[("selected", acc)])
        self.nb.configure(style="Rateio.TNotebook")

    def _sync_theme(self):
        self._style_notebook()
        for tab in (self.tab_geral, self.tab_horas, self.tab_custos,
                    self.tab_pct, self.tab_equip, self.tab_lanc, self.tab_resumo):
            tab.configure(fg_color=ThemeManager.p("surface"))
        if self.df_group is not None:
            self._populate_all()

    def _placeholder(self, msg=""):
        for tab in (self.tab_geral, self.tab_horas, self.tab_custos,
                    self.tab_pct, self.tab_equip, self.tab_lanc, self.tab_resumo):
            for w in tab.winfo_children():
                w.destroy()
            if msg:
                ctk.CTkLabel(tab, text=msg,
                             font=ctk.CTkFont("Segoe UI", 13),
                             text_color=ThemeManager.p("subtext")).pack(expand=True)

    # ── File Loading ───────────────────────────────────────────────────────────
    # ── File reading ───────────────────────────────────────────────────────────
    def _read_file(self, path):
        """
        Robustly read .xlsx / .xls / .csv (any separator & encoding) → DataFrame.
        Strategy for CSV:
          - Encodings tried: utf-8-sig (BOM), utf-8, latin-1, cp1252
          - Separators tried: auto-detect (python engine), ; , \\t
          - A result is accepted only when it produces ≥ 2 columns
        """
        ext = os.path.splitext(path)[1].lower()

        # ── Excel ──
        if ext in (".xlsx", ".xls", ".xlsm", ".ods"):
            return pd.read_excel(path)

        # ── CSV / TXT ──
        encodings  = ["utf-8-sig", "utf-8", "latin-1", "cp1252"]
        separators = [None, ";", ",", "\t"]   # None = pandas auto-detect

        last_err = None
        for enc in encodings:
            for sep in separators:
                try:
                    kwargs = {"encoding": enc}
                    if sep is None:
                        kwargs["sep"]    = None
                        kwargs["engine"] = "python"
                    else:
                        kwargs["sep"] = sep
                    # Ignore malformed lines instead of crashing
                    try:
                        kwargs["on_bad_lines"] = "skip"   # pandas ≥ 1.3
                        df = pd.read_csv(path, **kwargs)
                    except TypeError:
                        kwargs.pop("on_bad_lines", None)
                        kwargs["error_bad_lines"] = False  # pandas < 1.3
                        df = pd.read_csv(path, **kwargs)

                    # Accept only if we got a real table (≥ 2 columns)
                    if len(df.columns) >= 2:
                        return df
                except Exception as e:
                    last_err = e

        raise ValueError(
            f"Não foi possível interpretar o arquivo como tabela.\n"
            f"Encodings testados: {encodings}\n"
            f"Separadores testados: {[s or 'auto' for s in separators]}\n"
            f"Último erro: {last_err}"
        )

    def _normalise_df(self, df):
        """Strip column names and coerce numeric columns."""
        df.columns = [c.strip() for c in df.columns]
        for col in NUM_COLS:
            if col in df.columns:
                df[col] = pd.to_numeric(
                    df[col].astype(str)
                           .str.replace(",", ".", regex=False)
                           .str.replace(r"[^0-9.\-]", "", regex=True),
                    errors="coerce"
                )
        return df

    def _load_file(self):
        path = filedialog.askopenfilename(
            title="Selecionar arquivo de dados",
            filetypes=[
                ("Planilhas e CSV", "*.xlsx *.xls *.xlsm *.csv *.txt"),
                ("Excel",           "*.xlsx *.xls *.xlsm"),
                ("CSV / Texto",     "*.csv *.txt"),
                ("Todos",           "*.*"),
            ]
        )
        if not path:
            return
        try:
            df = self._read_file(path)
            df = self._normalise_df(df)

            if "CR" not in df.columns:
                messagebox.showerror(
                    "Coluna não encontrada",
                    "A coluna 'CR' não foi encontrada no arquivo.\n\n"
                    f"Colunas detectadas: {', '.join(df.columns.tolist())}"
                )
                return

            self.df_raw   = df
            self.df_group = self._build_groupby(df)
            self._db_pendencias = None # Limpa histórico do banco para priorizar o arquivo aberto
            self.lbl_file.configure(text=f"✓  {os.path.basename(path)}")
            self.export_btn.configure(state="normal")
            self._populate_all()

        except Exception as e:
            messagebox.showerror("Erro ao carregar arquivo", str(e))

    def _load_from_database(self):
        data_de_str = self.cal_de.get()
        data_para_str = self.cal_para.get()
        
        try:
            # Converte os dois pro formato do MySQL
            data_de_mysql = datetime.strptime(data_de_str, "%d/%m/%Y").strftime("%Y-%m-%d")
            data_para_mysql = datetime.strptime(data_para_str, "%d/%m/%Y").strftime("%Y-%m-%d")
        except ValueError:
            messagebox.showerror("Data Inválida", "Formato de data incorreto.")
            return
            
        # Garante que a data 'De' não é maior que a 'Para'
        if data_de_mysql > data_para_mysql:
            messagebox.showerror("Período Inválido", "A data 'De' não pode ser maior que a data 'Para'.")
            return
            
        try:
            # Puxa os dados somados do período!
            df_dia = buscar_lancamentos_por_periodo(data_de_mysql, data_para_mysql)
            df_pendentes = buscar_equipamentos_sem_lancamento(data_de_mysql, data_para_mysql)
            
            if df_dia.empty and df_pendentes.empty:
                messagebox.showinfo("Sem registros", f"Nenhum dado encontrado no banco entre {data_de_str} e {data_para_str}.")
                return
            
            self.df_raw = df_dia
            if not df_dia.empty:
                self.df_group = self._build_groupby(df_dia)
            else:
                self.df_group = pd.DataFrame(columns=["CR", "Qtd Equipamentos"])
            
            self._db_pendencias = df_pendentes
            
            # Mostra o período na interface
            texto_lbl = f"🗄️  Banco: {data_de_str}" if data_de_str == data_para_str else f"🗄️  Banco: {data_de_str} até {data_para_str}"
            self.lbl_file.configure(text=texto_lbl)
            self.export_btn.configure(state="normal")
            
            self._populate_all()
            
        except Exception as e:
            messagebox.showerror("Erro no Banco", f"Falha ao consultar o SGBD: {e}")

    # ── Export ─────────────────────────────────────────────────────────────────
    def _export_current_tab(self):
        if self.df_group is None:
            return

        g = self.df_group
        tab_index = self.nb.index(self.nb.select())
        tab_names = ["Geral", "Horas", "Custos", "Percentuais", "Equipamentos", "Lançamentos", "Resumo"]
        tab_name  = tab_names[tab_index]

        # Build the DataFrame and sheet config for each tab
        if tab_index == 0:  # Geral
            df_exp = g.copy()
            title  = "Visão Geral por CR"

        elif tab_index == 1:  # Horas
            cols = ["CR"] + [c for c in HR_COLS + ["Hrs Extras"] if c in g.columns]
            df_exp = g[cols].copy()
            title  = "Análise de Horas por CR"

        elif tab_index == 2:  # Custos
            cols = ["CR"] + [c for c in COST_COLS if c in g.columns]
            df_exp = g[cols].copy()
            title  = "Análise de Custos por CR"

        elif tab_index == 3:  # Percentuais
            cols = ["CR"] + [c for c in PCT_COLS if c in g.columns]
            df_exp = g[cols].copy()
            title  = "Percentuais Médios por CR"

        elif tab_index == 4:  # Equipamentos
            cols = ["CR"] + (["Qtd Equipamentos"] if "Qtd Equipamentos" in g.columns else [])
            df_exp = g[cols].sort_values("Qtd Equipamentos", ascending=False).copy() if "Qtd Equipamentos" in g.columns else g[["CR"]].copy()
            title  = "Equipamentos por CR"

        elif tab_index == 5:  # Lançamentos
            self._export_pendencias()
            return

        else:  # Resumo
            df_exp = g.copy()
            title  = "Resumo Executivo por CR"

        # Ask where to save
        default_name = f"Rateio_{tab_name}.xlsx"
        path = filedialog.asksaveasfilename(
            title="Salvar planilha",
            defaultextension=".xlsx",
            initialfile=default_name,
            filetypes=[("Excel", "*.xlsx")]
        )
        if not path:
            return

        try:
            self._export_df(df_exp, path, title, tab_name)
            messagebox.showinfo("Exportado com sucesso",
                                f"Arquivo salvo em:\n{path}")
        except Exception as e:
            messagebox.showerror("Erro ao exportar", str(e))

    def _export_df(self, df, path, title, sheet_name="Dados"):
        import openpyxl
        from openpyxl.styles import (Font, PatternFill, Alignment,
                                     Border, Side, numbers)
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name[:31]

        # ── Paleta ──
        COLOR_HEADER_BG = "1E2A45"   # azul escuro
        COLOR_HEADER_FG = "FFFFFF"
        COLOR_TITLE_BG  = "2D5BE3"   # azul accent
        COLOR_TITLE_FG  = "FFFFFF"
        COLOR_ALT_ROW   = "F0F4FF"   # linhas pares
        COLOR_BORDER    = "C5CDE0"

        thin = Side(style="thin", color=COLOR_BORDER)
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        n_cols = len(df.columns)

        # ── Linha 1: Título mesclado ──
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
        title_cell = ws.cell(row=1, column=1, value=title)
        title_cell.font      = Font(name="Segoe UI", bold=True, size=14, color=COLOR_TITLE_FG)
        title_cell.fill      = PatternFill("solid", fgColor=COLOR_TITLE_BG)
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 28

        # ── Linha 2: cabeçalho ──
        header_fill = PatternFill("solid", fgColor=COLOR_HEADER_BG)
        for col_idx, col_name in enumerate(df.columns, start=1):
            label = COL_MAP.get(col_name, col_name)
            cell  = ws.cell(row=2, column=col_idx, value=label)
            cell.font      = Font(name="Segoe UI", bold=True, size=11, color=COLOR_HEADER_FG)
            cell.fill      = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border    = border
        ws.row_dimensions[2].height = 22

        # ── Linhas de dados ──
        alt_fill = PatternFill("solid", fgColor=COLOR_ALT_ROW)

        # Number format helpers
        fmt_int   = '#,##0'
        fmt_float = '#,##0.00'
        fmt_brl   = 'R$ #,##0.00'
        fmt_pct   = '0.0"%"'

        for row_idx, (_, row_data) in enumerate(df.iterrows(), start=3):
            is_alt = (row_idx % 2 == 0)
            ws.row_dimensions[row_idx].height = 18
            for col_idx, col_name in enumerate(df.columns, start=1):
                raw = row_data[col_name]
                # Convert pandas NA
                try:
                    val = None if pd.isna(raw) else raw
                except Exception:
                    val = raw

                # Cast Int64 to plain int for openpyxl
                try:
                    import numpy as np
                    if isinstance(val, (np.integer,)):
                        val = int(val)
                    elif isinstance(val, (np.floating,)):
                        val = float(val)
                except Exception:
                    pass

                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.border    = border
                cell.alignment = Alignment(horizontal="center" if col_name != "CR" else "left",
                                           vertical="center")
                cell.font      = Font(name="Segoe UI", size=10)
                if is_alt:
                    cell.fill = alt_fill

                # Number formats
                if col_name in HR_COLS or col_name == "Hrs Extras" or col_name == "Qtd Equipamentos":
                    cell.number_format = fmt_int
                elif col_name in COST_COLS:
                    cell.number_format = fmt_brl
                elif col_name in PCT_COLS:
                    cell.number_format = fmt_pct
                elif isinstance(val, float):
                    cell.number_format = fmt_float

        # ── Largura automática por coluna ──
        sample_widths = df.astype(str).replace("<NA>", "").replace("nan", "").map(len).max()
        for col_idx, col_name in enumerate(df.columns, start=1):
            letter  = get_column_letter(col_idx)
            label   = COL_MAP.get(col_name, col_name)
            max_len = max(len(label), int(sample_widths.get(col_name, 0) or 0))
            ws.column_dimensions[letter].width = min(max(max_len + 3, 12), 38)

        # ── Linha de totais / médias ──
        total_row = len(df) + 3
        ws.row_dimensions[total_row].height = 20
        total_fill = PatternFill("solid", fgColor="1E2A45")
        for col_idx, col_name in enumerate(df.columns, start=1):
            cell = ws.cell(row=total_row, column=col_idx)
            cell.fill   = total_fill
            cell.border = border
            cell.font   = Font(name="Segoe UI", bold=True, size=10, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center", vertical="center")

            if col_name == "CR":
                cell.value = "TOTAL / MÉDIA"
                cell.alignment = Alignment(horizontal="left", vertical="center")
            elif col_name in HR_COLS or col_name == "Hrs Extras" or col_name == "Qtd Equipamentos":
                col_vals = df[col_name].dropna()
                cell.value = int(col_vals.sum())
                cell.number_format = fmt_int
            elif col_name in COST_COLS:
                cell.value = float(df[col_name].dropna().sum())
                cell.number_format = fmt_brl
            elif col_name in PCT_COLS:
                cell.value = float(df[col_name].dropna().mean())
                cell.number_format = fmt_pct

        # ── Freeze panes: cabeçalho fixo ──
        ws.freeze_panes = "A3"

        # ── Auto-filter ──
        ws.auto_filter.ref = f"A2:{get_column_letter(n_cols)}{len(df)+2}"

        wb.save(path)

    def _build_groupby(self, df):
        agg = {}
        for col in NUM_COLS:
            if col in df.columns:
                agg[col] = "mean" if col in PCT_COLS else "sum"
        if "Equipamento" in df.columns:
            agg["Equipamento"] = "count"

        g = df.groupby("CR").agg(agg).reset_index()
        g.rename(columns={"Equipamento": "Qtd Equipamentos"}, inplace=True)

        # Separar Hrs Disp positivas (disponíveis reais) e negativas (horas extras)
        if "Hrs Disp" in df.columns:
            pos = (df[df["Hrs Disp"] >= 0]
                   .groupby("CR")["Hrs Disp"].sum()
                   .reset_index().rename(columns={"Hrs Disp": "_pos"}))
            neg = (df[df["Hrs Disp"] < 0]
                   .groupby("CR")["Hrs Disp"].sum().abs()
                   .reset_index().rename(columns={"Hrs Disp": "Hrs Extras"}))
            g = g.merge(pos, on="CR", how="left").merge(neg, on="CR", how="left")
            g["Hrs Disp"] = g["_pos"].fillna(0).round(0).astype("Int64")
            g["Hrs Extras"] = g["Hrs Extras"].fillna(0).round(0).astype("Int64")
            g.drop(columns=["_pos"], inplace=True)

        # Demais horas → int
        for col in HR_COLS:
            if col in g.columns:
                g[col] = g[col].round(0).astype("Int64")

        return g

    # ── Populate all sub-tabs ──────────────────────────────────────────────────
    def _populate_all(self):
        g = self.df_group
        self._populate_geral(g)
        self._populate_horas(g)
        self._populate_custos(g)
        self._populate_pct(g)
        self._populate_equip(g)
        self._populate_lanc(self.df_raw)
        self._populate_resumo(g)

    # ── Tab: Geral ─────────────────────────────────────────────────────────────
    def _populate_geral(self, g):
        tab = self.tab_geral
        for w in tab.winfo_children(): w.destroy()

        ctk.CTkLabel(tab, text="Visão Geral por CR",
                     font=ctk.CTkFont("Segoe UI", 14, weight="bold"),
                     anchor="w").pack(fill="x", padx=20, pady=(16, 4))
        ctk.CTkLabel(tab, text=f"Total de CRs: {len(g)}   |   Total de equipamentos: {g['Qtd Equipamentos'].sum() if 'Qtd Equipamentos' in g else '—'}",
                     font=ctk.CTkFont("Segoe UI", 11),
                     text_color=ThemeManager.p("subtext"), anchor="w").pack(fill="x", padx=20, pady=(0, 10))

        self._make_table(tab, g)

    # ── Tab: Horas ─────────────────────────────────────────────────────────────
    def _populate_horas(self, g):
        tab = self.tab_horas
        for w in tab.winfo_children(): w.destroy()
        cols = [c for c in HR_COLS if c in g.columns]

        ctk.CTkLabel(tab, text="Análise de Horas por CR",
                     font=ctk.CTkFont("Segoe UI", 14, weight="bold"),
                     anchor="w").pack(fill="x", padx=20, pady=(16, 10))

        # ── Cards: horas normais ──
        stats_row = ctk.CTkFrame(tab, fg_color="transparent")
        stats_row.pack(fill="x", padx=20, pady=(0, 8))
        for i, col in enumerate(cols):
            card = ModernCard(stats_row)
            card.pack(side="left", expand=True, fill="both", padx=(0, 8) if i < len(cols)-1 else 0)
            ctk.CTkLabel(card, text=COL_MAP.get(col, col),
                         font=ctk.CTkFont("Segoe UI", 11),
                         text_color=ThemeManager.p("subtext")).pack(pady=(12, 2), padx=14, anchor="w")
            total = int(g[col].sum())
            avg   = int(g[col].mean())
            mx    = int(g[col].max())
            ctk.CTkLabel(card, text=f"{total:,} h",
                         font=ctk.CTkFont("Segoe UI", 20, weight="bold"),
                         text_color=CHART_COLORS[i]).pack(padx=14, anchor="w")
            ctk.CTkLabel(card, text=f"Média: {avg} h   |   Máx: {mx} h",
                         font=ctk.CTkFont("Segoe UI", 10),
                         text_color=ThemeManager.p("subtext")).pack(padx=14, pady=(0, 12), anchor="w")

        # ── Card: Horas Extras (horas negativas separadas) ──
        if "Hrs Extras" in g.columns and g["Hrs Extras"].sum() > 0:
            extras_row = ctk.CTkFrame(tab, fg_color="transparent")
            extras_row.pack(fill="x", padx=20, pady=(0, 10))

            header_card = ModernCard(extras_row, border_color=ThemeManager.p("danger"))
            header_card.pack(fill="x")
            inner = ctk.CTkFrame(header_card, fg_color="transparent")
            inner.pack(fill="x", padx=14, pady=10)
            inner.grid_columnconfigure(0, weight=1)

            ctk.CTkLabel(inner,
                         text="⚠️  Horas Extras  (originadas de Hrs Disponíveis negativas)",
                         font=ctk.CTkFont("Segoe UI", 11, weight="bold"),
                         text_color=ThemeManager.p("danger"),
                         anchor="w").grid(row=0, column=0, sticky="w", columnspan=4)

            # One mini-card per CR that has extras
            has_extras = g[g["Hrs Extras"] > 0].sort_values("Hrs Extras", ascending=False)
            sub_row = ctk.CTkFrame(inner, fg_color="transparent")
            sub_row.grid(row=1, column=0, sticky="ew", pady=(8, 0))
            for idx, (_, row_data) in enumerate(has_extras.iterrows()):
                mini = ctk.CTkFrame(sub_row, fg_color=ThemeManager.p("surface"),
                                    corner_radius=8, border_width=1,
                                    border_color=ThemeManager.p("border"))
                mini.pack(side="left", padx=(0, 6), pady=2)
                ctk.CTkLabel(mini, text=row_data["CR"][:20],
                             font=ctk.CTkFont("Segoe UI", 10),
                             text_color=ThemeManager.p("subtext")).pack(padx=10, pady=(6, 0), anchor="w")
                ctk.CTkLabel(mini, text=f"{int(row_data['Hrs Extras']):,} h",
                             font=ctk.CTkFont("Segoe UI", 14, weight="bold"),
                             text_color=ThemeManager.p("danger")).pack(padx=10, pady=(0, 6), anchor="w")

        # ── Charts side by side: horas normais | horas extras por CR ──
        chart_row = ctk.CTkFrame(tab, fg_color="transparent")
        chart_row.pack(fill="both", expand=True, padx=20, pady=(0, 10))
        chart_row.grid_columnconfigure(0, weight=3)

        if cols:
            fig = self._bar_chart(g, "CR", cols,
                                  [COL_MAP.get(c, c) for c in cols],
                                  "Horas por CR")
            f_main = ctk.CTkFrame(chart_row, fg_color="transparent")
            f_main.grid(row=0, column=0, sticky="nsew")
            self._embed_figure(f_main, fig, expand=True)

        if "Hrs Extras" in g.columns and g["Hrs Extras"].sum() > 0:
            chart_row.grid_columnconfigure(1, weight=1)
            fig_ex = self._bar_chart(g[g["Hrs Extras"] > 0].copy(),
                                     "CR", ["Hrs Extras"], ["Hrs Extras"],
                                     "Horas Extras por CR", start_color_idx=8)
            # force danger color
            f_ex = ctk.CTkFrame(chart_row, fg_color="transparent")
            f_ex.grid(row=0, column=1, sticky="nsew", padx=(8, 0))
            self._embed_figure(f_ex, fig_ex, expand=True)

    # ── Tab: Custos ────────────────────────────────────────────────────────────
    def _populate_custos(self, g):
        tab = self.tab_custos
        for w in tab.winfo_children(): w.destroy()
        cols = [c for c in COST_COLS if c in g.columns]

        ctk.CTkLabel(tab, text="Análise de Custos por CR",
                     font=ctk.CTkFont("Segoe UI", 14, weight="bold"),
                     anchor="w").pack(fill="x", padx=20, pady=(16, 10))

        stats_row = ctk.CTkFrame(tab, fg_color="transparent")
        stats_row.pack(fill="x", padx=20, pady=(0, 14))
        for i, col in enumerate(cols):
            card = ModernCard(stats_row)
            card.pack(side="left", expand=True, fill="both", padx=(0, 8) if i < len(cols)-1 else 0)
            ctk.CTkLabel(card, text=COL_MAP.get(col, col),
                         font=ctk.CTkFont("Segoe UI", 11),
                         text_color=ThemeManager.p("subtext")).pack(pady=(12, 2), padx=14, anchor="w")
            total = g[col].sum()
            avg   = g[col].mean()
            ctk.CTkLabel(card, text=f"R$ {total:,.2f}",
                         font=ctk.CTkFont("Segoe UI", 18, weight="bold"),
                         text_color=CHART_COLORS[i+3]).pack(padx=14, anchor="w")
            ctk.CTkLabel(card, text=f"Média por CR: R$ {avg:,.2f}",
                         font=ctk.CTkFont("Segoe UI", 10),
                         text_color=ThemeManager.p("subtext")).pack(padx=14, pady=(0, 12), anchor="w")

        if cols:
            fig = self._bar_chart(g, "CR", cols,
                                   [COL_MAP.get(c, c) for c in cols],
                                   "Custos por CR (R$)", start_color_idx=3)
            self._embed_figure(tab, fig, expand=True)

    # ── Tab: Percentuais ───────────────────────────────────────────────────────
    def _populate_pct(self, g):
        tab = self.tab_pct
        for w in tab.winfo_children(): w.destroy()
        cols = [c for c in PCT_COLS if c in g.columns]

        ctk.CTkLabel(tab, text="Análise de Percentuais por CR  (média por CR)",
                     font=ctk.CTkFont("Segoe UI", 14, weight="bold"),
                     anchor="w").pack(fill="x", padx=20, pady=(16, 10))

        stats_row = ctk.CTkFrame(tab, fg_color="transparent")
        stats_row.pack(fill="x", padx=20, pady=(0, 14))
        for i, col in enumerate(cols):
            card = ModernCard(stats_row)
            card.pack(side="left", expand=True, fill="both", padx=(0, 8) if i < len(cols)-1 else 0)
            ctk.CTkLabel(card, text=COL_MAP.get(col, col),
                         font=ctk.CTkFont("Segoe UI", 11),
                         text_color=ThemeManager.p("subtext")).pack(pady=(12, 2), padx=14, anchor="w")
            avg = g[col].mean()
            mn  = g[col].min()
            mx  = g[col].max()
            ctk.CTkLabel(card, text=f"{avg:.1f}%",
                         font=ctk.CTkFont("Segoe UI", 20, weight="bold"),
                         text_color=CHART_COLORS[i+6]).pack(padx=14, anchor="w")
            ctk.CTkLabel(card, text=f"Min: {mn:.1f}%   |   Máx: {mx:.1f}%",
                         font=ctk.CTkFont("Segoe UI", 10),
                         text_color=ThemeManager.p("subtext")).pack(padx=14, pady=(0, 12), anchor="w")

        if cols:
            fig = self._bar_chart(g, "CR", cols,
                                   [COL_MAP.get(c, c) for c in cols],
                                   "Percentuais Médios por CR (%)",
                                   start_color_idx=6, fmt_pct=True, ylim=(0, 100))
            self._embed_figure(tab, fig, expand=True)

    # ── Tab: Equipamentos ──────────────────────────────────────────────────────
    def _populate_equip(self, g):
        tab = self.tab_equip
        for w in tab.winfo_children(): w.destroy()

        ctk.CTkLabel(tab, text="Equipamentos por CR",
                     font=ctk.CTkFont("Segoe UI", 14, weight="bold"),
                     anchor="w").pack(fill="x", padx=20, pady=(16, 10))

        if "Qtd Equipamentos" not in g.columns:
            ctk.CTkLabel(tab, text="Coluna 'Equipamento' não encontrada.",
                         text_color=ThemeManager.p("subtext")).pack(expand=True)
            return

        # Ranking cards
        top = g.sort_values("Qtd Equipamentos", ascending=False).head(5)
        row = ctk.CTkFrame(tab, fg_color="transparent")
        row.pack(fill="x", padx=20, pady=(0, 14))
        total_eq = int(g["Qtd Equipamentos"].sum())
        for i, (_, r) in enumerate(top.iterrows()):
            card = ModernCard(row)
            card.pack(side="left", expand=True, fill="both", padx=(0, 6) if i < len(top)-1 else 0)
            pct = r["Qtd Equipamentos"] / total_eq * 100
            ctk.CTkLabel(card, text=r["CR"][:18],
                         font=ctk.CTkFont("Segoe UI", 11, weight="bold"),
                         text_color=ThemeManager.p("subtext")).pack(pady=(12, 2), padx=10, anchor="w")
            ctk.CTkLabel(card, text=str(int(r["Qtd Equipamentos"])),
                         font=ctk.CTkFont("Segoe UI", 22, weight="bold"),
                         text_color=CHART_COLORS[i]).pack(padx=10, anchor="w")
            ctk.CTkLabel(card, text=f"{pct:.1f}% do total",
                         font=ctk.CTkFont("Segoe UI", 10),
                         text_color=ThemeManager.p("subtext")).pack(padx=10, pady=(0, 12), anchor="w")

        self._make_table(tab, g[["CR", "Qtd Equipamentos"]].sort_values("Qtd Equipamentos", ascending=False))

    # ── Tab: Lançamentos ───────────────────────────────────────────────────────
    def _load_base_file(self):
        """Carrega a planilha base/gabarito para comparação de lançamentos."""
        path = filedialog.askopenfilename(
            title="Selecionar planilha BASE (gabarito de equipamentos)",
            filetypes=[
                ("Planilhas e CSV", "*.xlsx *.xls *.xlsm *.csv *.txt"),
                ("Excel",           "*.xlsx *.xls *.xlsm"),
                ("CSV / Texto",     "*.csv *.txt"),
                ("Todos",           "*.*"),
            ]
        )
        if not path:
            return
        try:
            df = self._read_file(path)
            df.columns = [c.strip() for c in df.columns]
            if "CR" not in df.columns or "Equipamento" not in df.columns:
                messagebox.showerror(
                    "Colunas não encontradas",
                    "A planilha base precisa ter as colunas 'CR' e 'Equipamento'.\n\n"
                    f"Colunas encontradas: {', '.join(df.columns.tolist())}"
                )
                return
            self.df_base = df[["CR", "Equipamento"]].drop_duplicates().reset_index(drop=True)
            self.lbl_base.configure(
                text=f"Base: {os.path.basename(path)} ({len(self.df_base)} equip.)",
                text_color=ThemeManager.p("warn")
            )
            # Atualiza a aba de lançamentos se já tiver dados carregados
            if self.df_raw is not None:
                self._populate_lanc(self.df_raw)
        except Exception as e:
            messagebox.showerror("Erro ao carregar base", str(e))

    def _get_sem_lancamento(self, df):
        """
        Retorna equipamentos sem lançamento.
        - Se df_base estiver carregado: compara a base com o df atual e retorna
          quem está na base mas NÃO aparece no df (não lançou no período).
        - Caso contrário: retorna linhas com horas zeradas (comportamento original).
        """
        # ── Modo comparação com base ──
        if self.df_base is not None and not self.df_base.empty:
            df_cols = df.copy()
            df_cols.columns = [c.strip() for c in df_cols.columns]

            # Normaliza CR e Equipamento para comparação case-insensitive
            base = self.df_base.copy()
            base["_cr_norm"]    = base["CR"].astype(str).str.strip().str.upper()
            base["_eq_norm"]    = base["Equipamento"].astype(str).str.strip().str.upper()

            if "CR" in df_cols.columns and "Equipamento" in df_cols.columns:
                lancados = df_cols[["CR", "Equipamento"]].dropna().drop_duplicates().copy()
                lancados["_cr_norm"] = lancados["CR"].astype(str).str.strip().str.upper()
                lancados["_eq_norm"] = lancados["Equipamento"].astype(str).str.strip().str.upper()
                encontrados = base.merge(
                    lancados[["_cr_norm", "_eq_norm"]].drop_duplicates(),
                    on=["_cr_norm", "_eq_norm"],
                    how="left",
                    indicator=True,
                )
                result = encontrados[encontrados["_merge"] == "left_only"][["CR", "Equipamento"]].copy()
            else:
                result = base[["CR", "Equipamento"]].copy()
            result["Status"] = "NÃ£o LanÃ§ado"
            return result.reset_index(drop=True)

            if "CR" in df_cols.columns and "Equipamento" in df_cols.columns:
                lançados = df_cols.copy()
                lançados["_cr_norm"] = lançados["CR"].astype(str).str.strip().str.upper()
                lançados["_eq_norm"] = lançados["Equipamento"].astype(str).str.strip().str.upper()
                chave_lançada = set(zip(lançados["_cr_norm"], lançados["_eq_norm"]))
            else:
                chave_lançada = set()

            # Quem está na base mas não foi lançado
            mask_faltando = ~base.apply(
                lambda r: (r["_cr_norm"], r["_eq_norm"]) in chave_lançada, axis=1
            )
            result = base[mask_faltando][["CR", "Equipamento"]].copy()
            result["Status"] = "Não Lançado"
            return result.reset_index(drop=True)

        # ── Modo original: horas zeradas ──
        check_cols = [c for c in ["Hrs Manu", "Hrs Trab", "Hrs Disp"] if c in df.columns]
        if not check_cols:
            return pd.DataFrame()
        mask = pd.Series([True] * len(df), index=df.index)
        for col in check_cols:
            mask = mask & (df[col].fillna(0).abs() < 0.01)
        result = df[mask].copy()
        display_cols = [c for c in
                        ["Equipamento", "CR", "Hrs Manu", "Hrs Trab", "Hrs Disp",
                         "$ Total Trab", "$ Total Disp", "% Util"]
                        if c in result.columns]
        return result[display_cols].reset_index(drop=True)

    

    def _export_pendencias(self):
        """Exporta planilha formatada de pendências por obra para cobrança."""
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from datetime import date

        # ── Coleta os dados ──
        if hasattr(self, "_db_pendencias") and self._db_pendencias is not None:
            sem_lanc = self._db_pendencias.copy()
        elif self.df_raw is not None:
            sem_lanc = self._get_sem_lancamento(self.df_raw)
        else:
            messagebox.showwarning("Sem dados", "Carregue um arquivo ou busque no banco antes de exportar.")
            return

        if sem_lanc.empty:
            messagebox.showinfo("Nenhuma pendência", "Não há equipamentos pendentes para exportar!")
            return

        # ── Escolhe onde salvar ──
        hoje = date.today().strftime("%d-%m-%Y")
        path = filedialog.asksaveasfilename(
            title="Salvar relatório de pendências",
            defaultextension=".xlsx",
            initialfile=f"Pendencias_Lancamento_{hoje}.xlsx",
            filetypes=[("Excel", "*.xlsx")]
        )
        if not path:
            return

        try:
            # ── Garante colunas mínimas ──
            if "CR" not in sem_lanc.columns:
                messagebox.showerror("Erro", "Os dados não possuem a coluna 'CR'.")
                return
            if "Equipamento" not in sem_lanc.columns:
                sem_lanc["Equipamento"] = "—"
            if "Status" not in sem_lanc.columns:
                sem_lanc["Status"] = "Pendente"

            # ── Agrupa por CR para gerar uma aba por obra ──
            wb = openpyxl.Workbook()
            wb.remove(wb.active)  # remove aba padrão vazia

            # Paleta
            C_TITLE_BG  = "C0392B"   # vermelho escuro – urgência
            C_TITLE_FG  = "FFFFFF"
            C_HEAD_BG   = "2C3E50"   # cinza azulado
            C_HEAD_FG   = "FFFFFF"
            C_ALT       = "FDECEA"   # vermelho bem claro para linhas pares
            C_BORDER    = "E0C0BE"
            C_TOTAL_BG  = "922B21"
            C_TOTAL_FG  = "FFFFFF"

            thin   = Side(style="thin",   color=C_BORDER)
            medium = Side(style="medium", color="922B21")
            border = Border(left=thin, right=thin, top=thin, bottom=thin)
            border_top = Border(left=thin, right=thin, top=medium, bottom=medium)

            alt_fill   = PatternFill("solid", fgColor=C_ALT)
            head_fill  = PatternFill("solid", fgColor=C_HEAD_BG)
            title_fill = PatternFill("solid", fgColor=C_TITLE_BG)
            total_fill = PatternFill("solid", fgColor=C_TOTAL_BG)

            grupos = sem_lanc.groupby("CR")
            crs_ordenados = sorted(grupos.groups.keys())

            # ── Aba resumo geral ──
            ws_res = wb.create_sheet("📋 Resumo Geral", 0)
            ws_res.sheet_view.showGridLines = False

            resumo_titulo = f"Relatório de Pendências de Lançamento — {hoje}"
            ws_res.merge_cells("A1:C1")
            tc = ws_res.cell(1, 1, resumo_titulo)
            tc.font      = Font("Segoe UI", bold=True, size=14, color=C_TITLE_FG)
            tc.fill      = title_fill
            tc.alignment = Alignment(horizontal="center", vertical="center")
            ws_res.row_dimensions[1].height = 32

            for col, label in enumerate(["Obra (CR)", "Equipamentos Pendentes", "Status"], 1):
                c = ws_res.cell(2, col, label)
                c.font      = Font("Segoe UI", bold=True, size=11, color=C_HEAD_FG)
                c.fill      = head_fill
                c.alignment = Alignment(horizontal="center", vertical="center")
                c.border    = border
            ws_res.row_dimensions[2].height = 22

            total_pendencias = 0
            for i, cr in enumerate(crs_ordenados, start=3):
                df_cr = grupos.get_group(cr)
                qtd   = len(df_cr)
                total_pendencias += qtd
                is_alt = (i % 2 == 0)
                for col, val in enumerate([cr, qtd, "⚠️ Pendente"], 1):
                    c = ws_res.cell(i, col, val)
                    c.font      = Font("Segoe UI", size=10)
                    c.alignment = Alignment(horizontal="center" if col > 1 else "left",
                                            vertical="center")
                    c.border    = border
                    if is_alt:
                        c.fill = alt_fill
                ws_res.row_dimensions[i].height = 18

            # Linha de total no resumo
            tr = len(crs_ordenados) + 3
            for col, val in enumerate(["TOTAL", total_pendencias, f"{len(crs_ordenados)} obras"], 1):
                c = ws_res.cell(tr, col, val)
                c.font      = Font("Segoe UI", bold=True, size=10, color=C_TOTAL_FG)
                c.fill      = total_fill
                c.alignment = Alignment(horizontal="center" if col > 1 else "left", vertical="center")
                c.border    = border_top
            ws_res.row_dimensions[tr].height = 20

            ws_res.column_dimensions["A"].width = 42
            ws_res.column_dimensions["B"].width = 24
            ws_res.column_dimensions["C"].width = 16
            ws_res.freeze_panes = "A3"

            # ── Uma aba por CR ──
            for cr in crs_ordenados:
                df_cr = grupos.get_group(cr).reset_index(drop=True)

                # Nome da aba: limita 31 chars (limite do Excel)
                sheet_name = str(cr)[:31]
                # Remove caracteres inválidos para nome de aba
                for ch in r"\/?*[]:'":
                    sheet_name = sheet_name.replace(ch, " ")
                ws = wb.create_sheet(sheet_name)
                ws.sheet_view.showGridLines = False

                # ── Linha 1: título da obra ──
                n_cols = max(len(df_cr.columns), 3)
                ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
                tc = ws.cell(1, 1, f"Equipamentos Pendentes — {cr}")
                tc.font      = Font("Segoe UI", bold=True, size=13, color=C_TITLE_FG)
                tc.fill      = title_fill
                tc.alignment = Alignment(horizontal="left", vertical="center")
                ws.row_dimensions[1].height = 28

                # ── Linha 2: subtítulo com data e total ──
                ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_cols)
                sub = ws.cell(2, 1,
                    f"Gerado em {hoje}  ·  {len(df_cr)} equipamento{'s' if len(df_cr) != 1 else ''} sem lançamento")
                sub.font      = Font("Segoe UI", italic=True, size=10, color="AAAAAA")
                sub.alignment = Alignment(horizontal="left", vertical="center")
                sub.fill      = PatternFill("solid", fgColor="1C1C1C")
                ws.row_dimensions[2].height = 18

                # ── Linha 3: cabeçalho de colunas ──
                col_labels = {"CR": "Obra (CR)", "Equipamento": "Equipamento",
                               "Status": "Status", "Hrs Manu": "Hrs Manutenção",
                               "Hrs Trab": "Hrs Trabalhadas", "Hrs Disp": "Hrs Disponíveis"}
                for ci, col in enumerate(df_cr.columns, 1):
                    c = ws.cell(3, ci, col_labels.get(col, col))
                    c.font      = Font("Segoe UI", bold=True, size=11, color=C_HEAD_FG)
                    c.fill      = head_fill
                    c.alignment = Alignment(horizontal="center", vertical="center")
                    c.border    = border
                ws.row_dimensions[3].height = 22

                # ── Dados ──
                for ri, (_, row) in enumerate(df_cr.iterrows(), start=4):
                    is_alt = (ri % 2 == 0)
                    for ci, col in enumerate(df_cr.columns, 1):
                        val = row[col]
                        try:
                            val = None if pd.isna(val) else val
                        except Exception:
                            pass
                        c = ws.cell(ri, ci, val)
                        c.font      = Font("Segoe UI", size=10)
                        c.alignment = Alignment(
                            horizontal="left" if col in ("CR", "Equipamento") else "center",
                            vertical="center")
                        c.border = border
                        if is_alt:
                            c.fill = alt_fill
                    ws.row_dimensions[ri].height = 18

                # ── Linha de total por aba ──
                tr_cr = len(df_cr) + 4
                ws.merge_cells(start_row=tr_cr, start_column=1, end_row=tr_cr, end_column=n_cols)
                c = ws.cell(tr_cr, 1, f"Total: {len(df_cr)} equipamento{'s' if len(df_cr) != 1 else ''} pendente{'s' if len(df_cr) != 1 else ''}")
                c.font      = Font("Segoe UI", bold=True, size=10, color=C_TOTAL_FG)
                c.fill      = total_fill
                c.alignment = Alignment(horizontal="left", vertical="center")
                c.border    = border_top
                ws.row_dimensions[tr_cr].height = 20

                # ── Largura automática ──
                for ci, col in enumerate(df_cr.columns, 1):
                    ltr     = get_column_letter(ci)
                    max_len = len(col_labels.get(col, col))
                    for val in df_cr[col].astype(str):
                        max_len = max(max_len, len(val))
                    ws.column_dimensions[ltr].width = min(max(max_len + 3, 14), 50)

                ws.freeze_panes = "A4"
                ws.auto_filter.ref = f"A3:{get_column_letter(len(df_cr.columns))}3"

            wb.save(path)
            messagebox.showinfo(
                "Exportado com sucesso",
                f"Relatório salvo em:\n{path}\n\n"
                f"📋 {len(crs_ordenados)} aba{'s' if len(crs_ordenados) != 1 else ''} de obra + 1 aba de resumo\n"
                f"⚠️  {total_pendencias} equipamento{'s' if total_pendencias != 1 else ''} pendente{'s' if total_pendencias != 1 else ''}"
            )
        except Exception as e:
            messagebox.showerror("Erro ao exportar pendências", str(e))

    def _populate_lanc(self, df_raw):
        tab = self.tab_lanc
        for w in tab.winfo_children(): w.destroy()

        # 1. Define a origem dos dados (Banco de Dados ou Filtro do Excel)
        if hasattr(self, "_db_pendencias") and self._db_pendencias is not None:
            sem_lanc = self._db_pendencias
        else:
            sem_lanc = self._get_sem_lancamento(df_raw)
            
        total_sem = len(sem_lanc)
        crs_sem   = sorted(sem_lanc["CR"].unique().tolist()) if "CR" in sem_lanc.columns and total_sem > 0 else []

        # 2. O resumo por CR continua aqui, usando o 'sem_lanc' atualizado!
        if total_sem > 0 and "CR" in sem_lanc.columns:
            cr_summary = (sem_lanc.groupby("CR")
                          .agg(Pendências=("Equipamento", "count"))
                          .reset_index()
                          .sort_values("Pendências", ascending=False))
        else:
            cr_summary = pd.DataFrame(columns=["CR", "Pendências"])

        # ── Estilização do Treeview ────────────────────────────────────────
        s = ttk.Style()
        s.configure("Lanc.Treeview",
                    background=ThemeManager.p("card"),
                    foreground=ThemeManager.p("text"),
                    fieldbackground=ThemeManager.p("card"),
                    rowheight=26, borderwidth=0,
                    font=("Segoe UI", 10))
        s.map("Lanc.Treeview", background=[("selected", ThemeManager.p("accent"))])
        s.configure("Lanc.Treeview.Heading",
                    background=ThemeManager.p("surface"),
                    foreground=ThemeManager.p("subtext"),
                    relief="flat", font=("Segoe UI", 10, "bold"))

        # Main container
        main = ctk.CTkFrame(tab, fg_color="transparent")
        main.pack(fill="both", expand=True, padx=12, pady=12)

        # ════════════════════════════════════════════════════════════════════
        # PAINEL ESQUERDO — Tabela de Resumo de CRs com Pendências
        # ════════════════════════════════════════════════════════════════════
        left_card = ModernCard(main)
        left_card.pack(side="left", fill="y", padx=(0, 6))
        left_card.configure(width=270)
        left_card.pack_propagate(False)

        lhdr = ctk.CTkFrame(left_card, fg_color="transparent")
        lhdr.pack(fill="x", padx=14, pady=(14, 6))
        ctk.CTkLabel(lhdr, text="🔔  Pendências por CR",
                     font=ctk.CTkFont("Segoe UI", 12, weight="bold"),
                     anchor="w").pack(side="left")

        badge_color = ThemeManager.p("danger") if total_sem > 0 else ThemeManager.p("success")
        ctk.CTkLabel(lhdr,
                     text=f"{len(crs_sem)} CR{'s' if len(crs_sem) != 1 else ''}",
                     font=ctk.CTkFont("Segoe UI", 10, weight="bold"),
                     text_color=badge_color).pack(side="right")

        sep = ctk.CTkFrame(left_card, height=1, fg_color=ThemeManager.p("border"))
        sep.pack(fill="x", padx=14, pady=(0, 6))

        btn_all = ctk.CTkButton(
            left_card,
            text=f"  📋  Todos  ({total_sem} equip.)",
            height=32, corner_radius=8,
            font=ctk.CTkFont("Segoe UI", 10, weight="bold"),
            fg_color=ThemeManager.p("accent"),
            hover_color=GlowButton._darken(ThemeManager.p("accent")),
            text_color="#ffffff",
            anchor="w",
        )
        btn_all.pack(fill="x", padx=10, pady=(0, 8))

        sum_frame = ctk.CTkFrame(left_card, fg_color="transparent")
        sum_frame.pack(fill="both", expand=True, padx=10, pady=(0, 10))
        sum_frame.grid_rowconfigure(0, weight=1)
        sum_frame.grid_columnconfigure(0, weight=1)

        sum_tree = ttk.Treeview(sum_frame, style="Lanc.Treeview",
                                columns=("cr", "pend"), show="headings",
                                selectmode="browse")
        sum_tree.heading("cr",   text="CR")
        sum_tree.heading("pend", text="Equipamentos")
        sum_tree.column("cr",   width=158, anchor="w")
        sum_tree.column("pend", width=80,  anchor="center")

        sum_vsb = ttk.Scrollbar(sum_frame, orient="vertical", command=sum_tree.yview)
        sum_tree.configure(yscrollcommand=sum_vsb.set)
        sum_tree.grid(row=0, column=0, sticky="nsew")
        sum_vsb.grid(row=0, column=1, sticky="ns")

        # AQUI a variável cr_summary sendo utilizada para preencher a interface!
        if cr_summary.empty:
            sum_tree.insert("", "end", values=("✓  Nenhuma pendência", ""), tags=("ok",))
            sum_tree.tag_configure("ok", foreground=ThemeManager.p("success"))
        else:
            for _, r in cr_summary.iterrows():
                sum_tree.insert("", "end",
                                values=(r["CR"], int(r["Pendências"])),
                                tags=("pend",), iid=r["CR"])
            sum_tree.tag_configure("pend", foreground=ThemeManager.p("warn"))

        # ════════════════════════════════════════════════════════════════════
        # PAINEL DIREITO — Detalhamento dos Equipamentos Sem Lançamento
        # ════════════════════════════════════════════════════════════════════
        right_card = ModernCard(main)
        right_card.pack(side="left", fill="both", expand=True, padx=(6, 0))
        right_card.grid_rowconfigure(1, weight=1)
        right_card.grid_columnconfigure(0, weight=1)

        rhdr = ctk.CTkFrame(right_card, fg_color="transparent")
        rhdr.pack(fill="x", padx=16, pady=(14, 6))
        rhdr.grid_columnconfigure(1, weight=1)

        self._lanc_title = ctk.CTkLabel(
            rhdr,
            text="Comparando com base: equipamentos não lançados" if self.df_base is not None
                 else "Todos os equipamentos sem lançamento (D-2)",
            font=ctk.CTkFont("Segoe UI", 12, weight="bold"), anchor="w")
        self._lanc_title.pack(side="left")

        # Botão exportar pendências
        GlowButton(rhdr, text="📤  Exportar Pendências",
                   command=self._export_pendencias,
                   color_key="danger", width=170, height=32).pack(side="right", padx=(8, 0))

        self._lanc_counter = ctk.CTkLabel(
            rhdr,
            text=f"{total_sem} equipamento{'s' if total_sem != 1 else ''}",
            font=ctk.CTkFont("Segoe UI", 11),
            text_color=ThemeManager.p("danger") if total_sem > 0 else ThemeManager.p("success"))
        self._lanc_counter.pack(side="right")

        sep2 = ctk.CTkFrame(right_card, height=1, fg_color=ThemeManager.p("border"))
        sep2.pack(fill="x", padx=16, pady=(0, 6))

        tree_frame = ctk.CTkFrame(right_card, fg_color="transparent")
        tree_frame.pack(fill="both", expand=True, padx=12, pady=(0, 12))
        tree_frame.grid_rowconfigure(0, weight=1)
        tree_frame.grid_columnconfigure(0, weight=1)

        tree_cols = list(sem_lanc.columns) if not sem_lanc.empty else \
                    ["Equipamento", "CR", "Hrs Manu", "Hrs Trab", "Hrs Disp"]

        self._lanc_tree = ttk.Treeview(tree_frame, style="Lanc.Treeview",
                                       columns=tree_cols, show="headings")
        col_widths = {"Equipamento": 110, "CR": 190, "Hrs Manu": 100,
                      "Hrs Trab": 110, "Hrs Disp": 110,
                      "$ Total Trab": 140, "$ Total Disp": 140, "% Util": 88}
        for col in tree_cols:
            self._lanc_tree.heading(col, text=COL_MAP.get(col, col),
                                    command=lambda c=col: self._lanc_sort(c))
            self._lanc_tree.column(col, width=col_widths.get(col, 110),
                                   anchor="w" if col in ("Equipamento", "CR") else "center")

        vsb = ttk.Scrollbar(tree_frame, orient="vertical",  command=self._lanc_tree.yview)
        hsb = ttk.Scrollbar(tree_frame, orient="horizontal", command=self._lanc_tree.xview)
        self._lanc_tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        self._lanc_tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")

        self._lanc_df      = sem_lanc
        self._lanc_sort_col = None
        self._lanc_sort_asc = True
        self._lanc_active_cr = "__ALL__"

        def _fmt_row(row, cols):
            vals = []
            for col in cols:
                v = row.get(col, "—") if col in row.index else "—"
                if v == "—":
                    vals.append("—")
                elif col in ["Hrs Manu", "Hrs Trab", "Hrs Disp"] and pd.notna(v):
                    try: vals.append(f"{int(float(v)):,}")
                    except ValueError: vals.append(str(v))
                elif col in COST_COLS and pd.notna(v):
                    try: vals.append(f"R$ {float(v):,.2f}")
                    except ValueError: vals.append(str(v))
                elif col in PCT_COLS and pd.notna(v):
                    try: vals.append(f"{float(v):.1f}%")
                    except ValueError: vals.append(str(v))
                else:
                    vals.append(str(v) if pd.notna(v) else "—")
            return vals

        def _refresh_detail(data, cr_label="Todos"):
            self._lanc_tree.delete(*self._lanc_tree.get_children())
            n = len(data)
            color = ThemeManager.p("danger") if n > 0 else ThemeManager.p("success")
            self._lanc_counter.configure(
                text=f"{n} equipamento{'s' if n != 1 else ''}",
                text_color=color)
            title = "Todos os equipamentos sem lançamento (D-2)" if cr_label == "Todos" \
                    else f"Sem lançamento  —  {cr_label}"
            if self.df_base is not None and cr_label == "Todos":
                title = "Comparando com base: equipamentos não lançados"
            elif self.df_base is not None:
                title = f"Não lançados (base vs período)  —  {cr_label}"
            self._lanc_title.configure(text=title)
            for _, row in data.iterrows():
                self._lanc_tree.insert("", "end",
                                       values=_fmt_row(row, tree_cols),
                                       tags=("pend",))
            self._lanc_tree.tag_configure("pend", foreground=ThemeManager.p("warn"))

        def _on_summary_select(event):
            sel = sum_tree.selection()
            if not sel:
                return
            cr = sel[0]
            self._lanc_active_cr = cr
            filtered = sem_lanc[sem_lanc["CR"] == cr] if "CR" in sem_lanc.columns else sem_lanc
            _refresh_detail(filtered, cr_label=cr)

        def _on_all():
            self._lanc_active_cr = "__ALL__"
            sum_tree.selection_remove(sum_tree.selection())
            _refresh_detail(sem_lanc)

        sum_tree.bind("<<TreeviewSelect>>", _on_summary_select)
        btn_all.configure(command=_on_all)

        _refresh_detail(sem_lanc)

    def _lanc_sort(self, col):
        """Sort the detail treeview by column header click."""
        if not hasattr(self, "_lanc_df") or self._lanc_df is None:
            return
        asc = not self._lanc_sort_asc if self._lanc_sort_col == col else True
        self._lanc_sort_col = col
        self._lanc_sort_asc = asc
        cr = getattr(self, "_lanc_active_cr", "__ALL__")
        data = self._lanc_df if cr == "__ALL__" else \
               self._lanc_df[self._lanc_df["CR"] == cr].copy()
        if col in data.columns:
            try:
                data = data.sort_values(col, ascending=asc)
            except Exception:
                pass
        self._lanc_tree.delete(*self._lanc_tree.get_children())
        tree_cols = [self._lanc_tree.heading(c)["text"] and c
                     for c in self._lanc_tree["columns"]]
        tree_cols = list(self._lanc_tree["columns"])
        for _, row in data.iterrows():
            vals = []
            for c in tree_cols:
                v = row.get(c, "—") if c in row.index else "—"
                if c in ["Hrs Manu", "Hrs Trab", "Hrs Disp"] and pd.notna(v):
                    vals.append(f"{int(float(v)):,}")
                elif c in COST_COLS and pd.notna(v):
                    vals.append(f"R$ {float(v):,.2f}")
                elif c in PCT_COLS and pd.notna(v):
                    vals.append(f"{float(v):.1f}%")
                else:
                    vals.append(str(v) if pd.notna(v) else "—")
            self._lanc_tree.insert("", "end", values=vals, tags=("pend",))
        self._lanc_tree.tag_configure("pend", foreground=ThemeManager.p("warn"))

    # ── Tab: Resumo ────────────────────────────────────────────────────────────
    def _populate_resumo(self, g):
        tab = self.tab_resumo
        for w in tab.winfo_children(): w.destroy()

        scroll_frame = ctk.CTkScrollableFrame(tab, fg_color="transparent")
        scroll_frame.pack(fill="both", expand=True)

        ctk.CTkLabel(scroll_frame, text="🗂️  Resumo Executivo",
                     font=ctk.CTkFont("Segoe UI", 16, weight="bold"),
                     anchor="w").pack(fill="x", padx=20, pady=(16, 4))
        ctk.CTkLabel(scroll_frame, text="Visão consolidada de todos os indicadores agrupados por CR.",
                     font=ctk.CTkFont("Segoe UI", 11),
                     text_color=ThemeManager.p("subtext"), anchor="w").pack(fill="x", padx=20, pady=(0, 14))

        # ── Gráfico 1: Horas (largura total, altura generosa) ──
        hr_cols = [c for c in HR_COLS if c in g.columns]
        if "Hrs Extras" in g.columns:
            hr_cols_chart = hr_cols + ["Hrs Extras"]
        else:
            hr_cols_chart = hr_cols

        if hr_cols_chart:
            ctk.CTkLabel(scroll_frame, text="⏱️  Horas por CR",
                         font=ctk.CTkFont("Segoe UI", 12, weight="bold"),
                         anchor="w").pack(fill="x", padx=20, pady=(0, 4))
            fig_h = self._bar_chart(g, "CR", hr_cols_chart,
                                    [COL_MAP.get(c, c) for c in hr_cols_chart],
                                    "", height=4.2, width=12)
            card_h = ModernCard(scroll_frame)
            card_h.pack(fill="x", padx=20, pady=(0, 14))
            self._embed_figure(card_h, fig_h)

        # ── Gráfico 2: Custos (largura total) ──
        cost_cols = [c for c in COST_COLS if c in g.columns]
        if cost_cols:
            ctk.CTkLabel(scroll_frame, text="💰  Custos por CR (R$)",
                         font=ctk.CTkFont("Segoe UI", 12, weight="bold"),
                         anchor="w").pack(fill="x", padx=20, pady=(0, 4))
            fig_c = self._bar_chart(g, "CR", cost_cols,
                                    [COL_MAP.get(c, c) for c in cost_cols],
                                    "", start_color_idx=3, height=4.2, width=12)
            card_c = ModernCard(scroll_frame)
            card_c.pack(fill="x", padx=20, pady=(0, 14))
            self._embed_figure(card_c, fig_c)

        # ── Gráfico 3: Percentuais (largura total, ylim 0-100) ──
        pct_cols = [c for c in PCT_COLS if c in g.columns]
        if pct_cols:
            ctk.CTkLabel(scroll_frame, text="📈  Percentuais Médios por CR (%)",
                         font=ctk.CTkFont("Segoe UI", 12, weight="bold"),
                         anchor="w").pack(fill="x", padx=20, pady=(0, 4))
            fig_pct = self._bar_chart(g, "CR", pct_cols,
                                      [COL_MAP.get(c, c) for c in pct_cols],
                                      "", start_color_idx=6,
                                      fmt_pct=True, ylim=(0, 100),
                                      height=4.2, width=12)
            card_pct = ModernCard(scroll_frame)
            card_pct.pack(fill="x", padx=20, pady=(0, 14))
            self._embed_figure(card_pct, fig_pct)

        # ── Compare button ──
        sep = ctk.CTkFrame(scroll_frame, height=1, fg_color=ThemeManager.p("border"))
        sep.pack(fill="x", padx=20, pady=16)

        cmp_row = ctk.CTkFrame(scroll_frame, fg_color="transparent")
        cmp_row.pack(fill="x", padx=20, pady=(0, 20))
        ctk.CTkLabel(cmp_row,
                     text="🔄  Comparar com outro período:",
                     font=ctk.CTkFont("Segoe UI", 13, weight="bold")).pack(side="left")
        GlowButton(cmp_row, text="📂  Selecionar arquivo para comparativo",
                   command=self._load_compare,
                   color_key="accent2", width=280, height=36).pack(side="right")

        if self.df_compare_group is not None:
            self._build_compare_section(scroll_frame)

    # ── Compare ────────────────────────────────────────────────────────────────
    def _load_compare(self):
        path = filedialog.askopenfilename(
            title="Selecionar arquivo para comparativo",
            filetypes=[
                ("Planilhas e CSV", "*.xlsx *.xls *.xlsm *.csv *.txt"),
                ("Excel",           "*.xlsx *.xls *.xlsm"),
                ("CSV / Texto",     "*.csv *.txt"),
                ("Todos",           "*.*"),
            ]
        )
        if not path:
            return
        try:
            df = self._read_file(path)
            df = self._normalise_df(df)

            if "CR" not in df.columns:
                messagebox.showerror(
                    "Coluna não encontrada",
                    "A coluna 'CR' não foi encontrada no arquivo de comparação.\n\n"
                    f"Colunas detectadas: {', '.join(df.columns.tolist())}"
                )
                return

            self.df_compare       = df
            self.df_compare_group = self._build_groupby(df)
            self._populate_resumo(self.df_group)

        except Exception as e:
            messagebox.showerror("Erro ao carregar arquivo de comparação", str(e))

    def _build_compare_section(self, parent):
        g1 = self.df_group
        g2 = self.df_compare_group

        ctk.CTkLabel(parent, text="📊  Comparativo entre Períodos",
                     font=ctk.CTkFont("Segoe UI", 14, weight="bold"),
                     anchor="w").pack(fill="x", padx=20, pady=(8, 4))

        merged = g1.merge(g2, on="CR", suffixes=(" (A)", " (B)"), how="outer").fillna(0)

        # Delta table
        delta_data = {"CR": merged["CR"].tolist()}
        for col in HR_COLS + COST_COLS:
            ca, cb = f"{col} (A)", f"{col} (B)"
            if ca in merged.columns and cb in merged.columns:
                delta_data[f"Δ {COL_MAP.get(col, col)}"] = (merged[cb] - merged[ca]).round(1).tolist()

        delta_df = pd.DataFrame(delta_data)

        row = ctk.CTkFrame(parent, fg_color="transparent")
        row.pack(fill="x", padx=20, pady=(0, 14))
        row.grid_columnconfigure((0, 1), weight=1)

        hr_cols = [c for c in HR_COLS if c in g1.columns and c in g2.columns]
        if hr_cols:
            fig = self._compare_bar(merged, "CR", hr_cols,
                                    [COL_MAP.get(c, c) for c in hr_cols],
                                    "Comparativo – Horas")
            f = ctk.CTkFrame(row, fg_color="transparent")
            f.grid(row=0, column=0, sticky="nsew", padx=(0, 6))
            self._embed_figure(f, fig)

        cost_cols = [c for c in COST_COLS if c in g1.columns and c in g2.columns]
        if cost_cols:
            fig2 = self._compare_bar(merged, "CR", cost_cols,
                                     [COL_MAP.get(c, c) for c in cost_cols],
                                     "Comparativo – Custos (R$)", start_color_idx=3)
            f2 = ctk.CTkFrame(row, fg_color="transparent")
            f2.grid(row=0, column=1, sticky="nsew", padx=(6, 0))
            self._embed_figure(f2, fig2)

        ctk.CTkLabel(parent, text="Tabela de Deltas (B − A)",
                     font=ctk.CTkFont("Segoe UI", 13, weight="bold"),
                     anchor="w").pack(fill="x", padx=20, pady=(8, 4))
        self._make_table(parent, delta_df, height=min(len(delta_df) + 1, 8))

    # ── Chart helpers ──────────────────────────────────────────────────────────
    def _mpl_fig(self, width=10, height=3.5):
        bg  = ThemeManager.p("mpl_bg")
        fig = Figure(figsize=(width, height), facecolor=bg, tight_layout=True)
        return fig

    def _style_ax(self, ax, title=""):
        bg   = ThemeManager.p("mpl_bg")
        grid = ThemeManager.p("mpl_grid")
        txt  = ThemeManager.p("mpl_text")
        sub  = ThemeManager.p("subtext")
        ax.set_facecolor(bg)
        ax.set_title(title, color=txt, fontsize=11, fontweight="bold", pad=8)
        ax.tick_params(colors=sub, labelsize=8)
        for spine in ax.spines.values():
            spine.set_edgecolor(grid)
        ax.yaxis.grid(True, color=grid, linewidth=0.5, linestyle="--")
        ax.set_axisbelow(True)

    def _bar_chart(self, g, x_col, y_cols, labels, title, start_color_idx=0,
                   fmt_pct=False, height=3.5, width=10, ylim=None):
        fig = self._mpl_fig(width, height)
        ax  = fig.add_subplot(111)
        self._style_ax(ax, title)

        n    = len(g)
        nc   = len(y_cols)
        bw   = 0.7 / nc
        xs   = np.arange(n)

        for i, (col, lbl) in enumerate(zip(y_cols, labels)):
            if col not in g.columns:
                continue
            vals = g[col].fillna(0).tolist()
            offset = (i - nc/2 + 0.5) * bw
            bars = ax.bar(xs + offset, vals, width=bw,
                          color=CHART_COLORS[(start_color_idx + i) % len(CHART_COLORS)],
                          label=lbl, zorder=3)
            for bar, v in zip(bars, vals):
                if v != 0:
                    txt = f"{v:.0f}%" if fmt_pct else (f"{int(v)}" if col in HR_COLS else f"R$ {v:,.0f}")
                    ax.text(bar.get_x() + bar.get_width()/2,
                            bar.get_height() + max(vals)*0.01,
                            txt, ha="center", va="bottom", fontsize=7,
                            color=ThemeManager.p("mpl_text"))

        ax.set_xticks(xs)
        ax.set_xticklabels(g[x_col].tolist(), rotation=25, ha="right", fontsize=8)
        if ylim:
            ax.set_ylim(ylim)
        ax.legend(fontsize=8, facecolor=ThemeManager.p("card"),
                  labelcolor=ThemeManager.p("mpl_text"), edgecolor=ThemeManager.p("border"))
        return fig

    def _pie_chart(self, g, label_col, val_col, title):
        if val_col not in g.columns:
            return self._mpl_fig()
        fig = self._mpl_fig(5, 3.5)
        ax  = fig.add_subplot(111)
        ax.set_facecolor(ThemeManager.p("mpl_bg"))
        fig.patch.set_facecolor(ThemeManager.p("mpl_bg"))

        vals   = g[val_col].fillna(0).tolist()
        labels = g[label_col].tolist()
        colors = [CHART_COLORS[i % len(CHART_COLORS)] for i in range(len(labels))]

        wedges, texts, autotexts = ax.pie(
            vals, labels=None, autopct="%1.1f%%",
            colors=colors, startangle=140,
            pctdistance=0.78,
            wedgeprops=dict(edgecolor=ThemeManager.p("mpl_bg"), linewidth=2)
        )
        for t in autotexts:
            t.set_color(ThemeManager.p("mpl_bg"))
            t.set_fontsize(8)
            t.set_fontweight("bold")

        ax.legend(wedges, labels, loc="center left", bbox_to_anchor=(1, 0.5),
                  fontsize=8, facecolor=ThemeManager.p("card"),
                  labelcolor=ThemeManager.p("mpl_text"),
                  edgecolor=ThemeManager.p("border"))
        ax.set_title(title, color=ThemeManager.p("mpl_text"), fontsize=10, fontweight="bold")
        return fig

    def _compare_bar(self, merged, cr_col, cols, labels, title, start_color_idx=0):
        fig = self._mpl_fig(10, 3.5)
        ax  = fig.add_subplot(111)
        self._style_ax(ax, title)

        n  = len(merged)
        xs = np.arange(n)
        bw = 0.35

        for i, (col, lbl) in enumerate(zip(cols, labels)):
            ca, cb = f"{col} (A)", f"{col} (B)"
            if ca not in merged.columns or cb not in merged.columns:
                continue
            va = merged[ca].fillna(0).tolist()
            vb = merged[cb].fillna(0).tolist()
            c  = CHART_COLORS[(start_color_idx + i*2) % len(CHART_COLORS)]
            c2 = CHART_COLORS[(start_color_idx + i*2 + 1) % len(CHART_COLORS)]
            ax.bar(xs - bw/2, va, bw, color=c, label=f"{lbl} (A)", zorder=3)
            ax.bar(xs + bw/2, vb, bw, color=c2, label=f"{lbl} (B)", zorder=3, alpha=0.85)

        ax.set_xticks(xs)
        ax.set_xticklabels(merged[cr_col].tolist(), rotation=25, ha="right", fontsize=8)
        ax.legend(fontsize=8, facecolor=ThemeManager.p("card"),
                  labelcolor=ThemeManager.p("mpl_text"), edgecolor=ThemeManager.p("border"))
        return fig

    # ── Embed figure ──────────────────────────────────────────────────────────
    def _embed_figure(self, parent, fig, expand=False):
        canvas = FigureCanvasTkAgg(fig, master=parent)
        canvas.draw()
        canvas.get_tk_widget().pack(fill="both", expand=expand, padx=4, pady=4)
        self._charts.append(canvas)
        return canvas

    # ── Treeview table ────────────────────────────────────────────────────────
    def _make_table(self, parent, df, height=None):
        s = ttk.Style()
        s.configure("Rateio.Treeview",
                    background=ThemeManager.p("card"),
                    foreground=ThemeManager.p("text"),
                    fieldbackground=ThemeManager.p("card"),
                    rowheight=26, borderwidth=0,
                    font=("Segoe UI", 10))
        s.map("Rateio.Treeview", background=[("selected", ThemeManager.p("accent"))])
        s.configure("Rateio.Treeview.Heading",
                    background=ThemeManager.p("surface"),
                    foreground=ThemeManager.p("subtext"),
                    relief="flat", font=("Segoe UI", 10, "bold"))

        fr = ctk.CTkFrame(parent, fg_color="transparent")
        fr.pack(fill="both", expand=True, padx=20, pady=(0, 10))

        cols = list(df.columns)
        tv   = ttk.Treeview(fr, style="Rateio.Treeview",
                             columns=cols, show="headings",
                             height=height or min(len(df) + 1, 12))
        for col in cols:
            tv.heading(col, text=COL_MAP.get(col, col))
            tv.column(col, anchor="center", width=max(80, len(COL_MAP.get(col, col)) * 9))

        for _, row in df.iterrows():
            vals = []
            for col in cols:
                v = row[col]
                if col in HR_COLS:
                    vals.append(f"{int(v):,}" if pd.notna(v) else "—")
                elif col in COST_COLS:
                    vals.append(f"R$ {v:,.2f}" if pd.notna(v) else "—")
                elif col in PCT_COLS:
                    vals.append(f"{v:.1f}%" if pd.notna(v) else "—")
                else:
                    vals.append(str(v) if pd.notna(v) else "—")
            tv.insert("", "end", values=vals)

        sb = ttk.Scrollbar(fr, orient="vertical", command=tv.yview)
        tv.configure(yscrollcommand=sb.set)
        sbx = ttk.Scrollbar(fr, orient="horizontal", command=tv.xview)
        tv.configure(xscrollcommand=sbx.set)
        tv.grid(row=0, column=0, sticky="nsew")
        sb.grid(row=0, column=1, sticky="ns")
        sbx.grid(row=1, column=0, sticky="ew")
        fr.grid_rowconfigure(0, weight=1)
        fr.grid_columnconfigure(0, weight=1)


# ─── Main App ──────────────────────────────────────────────────────────────────

class App(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title("Gerador de Relatórios Automatizados")
        self.geometry("1100x700")
        self.minsize(900, 600)
        ctk.set_appearance_mode("Dark")
        ctk.set_default_color_theme("blue")

        self.grid_rowconfigure(1, weight=1)
        self.grid_columnconfigure(0, weight=1)

        self._build_header()
        self._build_tabs()
        ThemeManager.register(self._sync_header)

    # ── Header ─────────────────────────────────────────────────────────────────
    def _build_header(self):
        self.header = ctk.CTkFrame(self, height=58, corner_radius=0,
                                   fg_color=ThemeManager.p("surface"),
                                   border_width=0)
        self.header.grid(row=0, column=0, sticky="ew")
        self.header.grid_propagate(False)
        self.header.grid_columnconfigure(1, weight=1)

        # Logo / title
        ctk.CTkLabel(self.header,
                     text="  ⚙️  RelatórioBot",
                     font=ctk.CTkFont("Segoe UI", 16, weight="bold"),
                     text_color=ThemeManager.p("accent")).grid(row=0, column=0, padx=18, sticky="w")

        # Subtitle
        ctk.CTkLabel(self.header,
                     text="Automação · Análise · Rateio",
                     font=ctk.CTkFont("Segoe UI", 10),
                     text_color=ThemeManager.p("subtext")).grid(row=0, column=1, sticky="w")

        # Theme toggle
        self.theme_btn = ctk.CTkButton(
            self.header,
            text="☀️  Modo Claro",
            width=130, height=32, corner_radius=8,
            font=ctk.CTkFont("Segoe UI", 11),
            fg_color=ThemeManager.p("card"),
            hover_color=ThemeManager.p("border"),
            text_color=ThemeManager.p("text"),
            command=self._toggle_theme
        )
        self.theme_btn.grid(row=0, column=2, padx=18, sticky="e")

    def _sync_header(self):
        self.header.configure(fg_color=ThemeManager.p("surface"))
        label = "🌙  Modo Escuro" if ThemeManager.is_dark() else "☀️  Modo Claro"
        self.theme_btn.configure(
            text=label,
            fg_color=ThemeManager.p("card"),
            hover_color=ThemeManager.p("border"),
            text_color=ThemeManager.p("text")
        )

    def _toggle_theme(self):
        ThemeManager.toggle()

    # ── Tabs ───────────────────────────────────────────────────────────────────
    def _build_tabs(self):
        self.tab_view = ctk.CTkTabview(self, corner_radius=12,
                                       fg_color=ThemeManager.p("bg"),
                                       segmented_button_fg_color=ThemeManager.p("surface"),
                                       segmented_button_selected_color=ThemeManager.p("accent"),
                                       segmented_button_unselected_color=ThemeManager.p("surface"),
                                       segmented_button_selected_hover_color=ThemeManager.p("accent"),
                                       text_color=ThemeManager.p("text"),
                                       text_color_disabled=ThemeManager.p("subtext"))
        self.tab_view.grid(row=1, column=0, sticky="nsew", padx=16, pady=(8, 16))

        self.tab_view.add("  🤖  Bot  ")
        self.tab_view.add("  📊  Rateio  ")

        self.tab_view.tab("  🤖  Bot  ").grid_rowconfigure(0, weight=1)
        self.tab_view.tab("  🤖  Bot  ").grid_columnconfigure(0, weight=1)
        self.tab_view.tab("  📊  Rateio  ").grid_rowconfigure(0, weight=1)
        self.tab_view.tab("  📊  Rateio  ").grid_columnconfigure(0, weight=1)

        BotTab(self.tab_view.tab("  🤖  Bot  ")).grid(sticky="nsew")
        RateioTab(self.tab_view.tab("  📊  Rateio  ")).grid(sticky="nsew")
