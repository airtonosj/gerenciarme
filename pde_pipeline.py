from __future__ import annotations

import re
import unicodedata
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Iterable

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


NUMERIC_COLUMNS = [
    "Hrs Manu",
    "Hrs Trab",
    "Hrs Disp",
    "$ Total Trab",
    "$ Total Disp",
    "% Disp",
    "% Util",
    "MTBF",
    "MTTR",
]


CANONICAL_COLUMNS = {
    "equipamento": "Equipamento",
    "cr": "CR",
    "obra": "CR",
    "centro de resultado": "CR",
    "data(de)": "Data(de)",
    "data de": "Data(de)",
    "data inicio": "Data(de)",
    "data_inicio": "Data(de)",
    "data(para)": "Data(para)",
    "data para": "Data(para)",
    "data fim": "Data(para)",
    "data_fim": "Data(para)",
    "hrs manu": "Hrs Manu",
    "hrs manutencao": "Hrs Manu",
    "hrs trab": "Hrs Trab",
    "hrs trabalhadas": "Hrs Trab",
    "hrs disp": "Hrs Disp",
    "hrs disponiveis": "Hrs Disp",
    "$ total trab": "$ Total Trab",
    "total trab": "$ Total Trab",
    "$ total disp": "$ Total Disp",
    "total disp": "$ Total Disp",
    "% disp": "% Disp",
    "% util": "% Util",
    "mtbf": "MTBF",
    "mttr": "MTTR",
}


@dataclass(frozen=True)
class PDEResult:
    dashboard: pd.DataFrame
    pendencias: pd.DataFrame
    pendencias_por_obra: pd.DataFrame
    ativos_lancaram: pd.DataFrame
    validar_antes: pd.DataFrame
    transferencias: pd.DataFrame
    fora_base_ativa: pd.DataFrame
    duplicidades_atual: pd.DataFrame
    resumo_atual: pd.DataFrame
    base_mestre: pd.DataFrame
    historico_cr: pd.DataFrame


def normalize_text(value) -> str:
    if pd.isna(value):
        return ""
    text = str(value).strip()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = re.sub(r"\s+", " ", text)
    return text.upper()


def normalize_column(value) -> str:
    text = normalize_text(value).lower()
    text = text.replace("_", " ")
    text = re.sub(r"\s+", " ", text).strip()
    return CANONICAL_COLUMNS.get(text, str(value).strip())


def read_table(path: str | Path, sheet_name: str | int | None = None) -> pd.DataFrame:
    path = Path(path)
    suffix = path.suffix.lower()
    if suffix in {".xlsx", ".xls", ".xlsm", ".ods"}:
        return pd.read_excel(path, sheet_name=0 if sheet_name is None else sheet_name)

    encodings = ["utf-8-sig", "utf-8", "latin-1", "cp1252"]
    separators = [None, ";", ",", "\t"]
    last_error: Exception | None = None
    for encoding in encodings:
        for sep in separators:
            try:
                kwargs = {"encoding": encoding, "on_bad_lines": "skip"}
                if sep is None:
                    kwargs.update({"sep": None, "engine": "python"})
                else:
                    kwargs["sep"] = sep
                df = pd.read_csv(path, **kwargs)
                if len(df.columns) >= 2:
                    return df
            except Exception as exc:
                last_error = exc
    raise ValueError(f"Nao foi possivel ler {path}. Ultimo erro: {last_error}")


def find_master_sheet(path: str | Path) -> pd.DataFrame:
    path = Path(path)
    if path.suffix.lower() not in {".xlsx", ".xls", ".xlsm", ".ods"}:
        return read_table(path)

    excel = pd.ExcelFile(path)
    best_df: pd.DataFrame | None = None
    best_score = -1
    for sheet in excel.sheet_names:
        df = pd.read_excel(path, sheet_name=sheet)
        normalized = {normalize_column(col) for col in df.columns}
        score = 0
        for required in ["Equipamento", "CR", "CR provavel / validar", "Cobra PDE?"]:
            if required in normalized:
                score += 1
        if "Equipamento" in normalized and score > best_score:
            best_score = score
            best_df = df
    if best_df is None:
        raise ValueError(f"Nenhuma aba com coluna Equipamento encontrada em {path}")
    return best_df


def coerce_number(value) -> float:
    if pd.isna(value) or value == "":
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)

    text = str(value).strip()
    if text in {"-", "--", "—"}:
        return 0.0
    text = re.sub(r"[^0-9,\.\-]", "", text)
    if "." in text and "," in text:
        text = text.replace(".", "").replace(",", ".")
    elif "," in text:
        text = text.replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return 0.0


def coerce_date(value):
    if pd.isna(value) or value == "":
        return pd.NaT
    parsed = pd.to_datetime(value, dayfirst=True, errors="coerce")
    return parsed


def prepare_report(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    out.columns = [normalize_column(col) for col in out.columns]

    if "Equipamento" not in out.columns:
        raise ValueError("A coluna 'Equipamento' nao foi encontrada.")
    if "CR" not in out.columns:
        raise ValueError("A coluna 'CR' nao foi encontrada.")

    out["Equipamento"] = out["Equipamento"].astype(str).str.strip()
    out["CR"] = out["CR"].astype(str).str.strip()
    out = out[(out["Equipamento"] != "") & (out["CR"] != "")]

    for col in NUMERIC_COLUMNS:
        if col in out.columns:
            out[col] = out[col].map(coerce_number)
        else:
            out[col] = 0.0

    for col in ["Data(de)", "Data(para)"]:
        if col in out.columns:
            out[col] = out[col].map(coerce_date)
        else:
            out[col] = pd.NaT

    out["_eq_key"] = out["Equipamento"].map(normalize_text)
    out["_cr_key"] = out["CR"].map(normalize_text)
    return out


def read_reports(paths: Iterable[str | Path]) -> pd.DataFrame:
    frames = [prepare_report(read_table(path)) for path in paths]
    if not frames:
        raise ValueError("Informe ao menos um relatorio.")
    return pd.concat(frames, ignore_index=True)


def _first_non_empty(series: pd.Series) -> str:
    for value in series:
        if pd.notna(value) and str(value).strip():
            return str(value).strip()
    return ""


def build_master_base(historico: pd.DataFrame, reference_date=None) -> tuple[pd.DataFrame, pd.DataFrame]:
    hist = prepare_report(historico)
    if reference_date is None:
        reference_date = hist["Data(para)"].max()
    reference_date = pd.to_datetime(reference_date, dayfirst=True, errors="coerce")

    historico_cr = (
        hist.groupby(["_eq_key", "_cr_key"], as_index=False)
        .agg(
            Equipamento=("Equipamento", _first_non_empty),
            CR=("CR", _first_non_empty),
            **{col: (col, "sum") for col in NUMERIC_COLUMNS},
            Primeira_aparicao=("Data(de)", "min"),
            Ultima_aparicao=("Data(para)", "max"),
            Qtd_linhas=("Equipamento", "count"),
        )
        .sort_values(["Equipamento", "CR"])
    )

    candidates = historico_cr.sort_values(
        ["_eq_key", "Ultima_aparicao", "Hrs Trab", "Hrs Disp"],
        ascending=[True, False, False, False],
    )
    last_seen = candidates.drop_duplicates("_eq_key", keep="first").set_index("_eq_key")

    grouped = historico_cr.groupby("_eq_key")
    master = grouped.agg(
        Equipamento=("Equipamento", _first_non_empty),
        Qtd_CRs_historico=("CR", "nunique"),
        CRs_onde_apareceu=("CR", lambda s: " | ".join(sorted(set(map(str, s))))),
        Total_Hrs_Trab=("Hrs Trab", "sum"),
        Total_Hrs_Disp=("Hrs Disp", "sum"),
        Total_Hrs_Manu=("Hrs Manu", "sum"),
        Primeira_aparicao=("Primeira_aparicao", "min"),
        Ultima_aparicao=("Ultima_aparicao", "max"),
    ).reset_index()

    master["CR provavel / validar"] = master["_eq_key"].map(last_seen["CR"])
    master["Status sugerido"] = master["Qtd_CRs_historico"].map(
        lambda qtd: "CANDIDATO ATIVO" if int(qtd) == 1 else "VALIDAR CR ATUAL"
    )
    master["Cobra PDE?"] = master["Qtd_CRs_historico"].map(
        lambda qtd: "SIM" if int(qtd) == 1 else "VALIDAR"
    )
    master["Status manual"] = ""
    master["Observacao"] = ""
    master["Dias sem lancamento"] = (
        reference_date.normalize() - pd.to_datetime(master["Ultima_aparicao"]).dt.normalize()
    ).dt.days

    ordered = [
        "Equipamento",
        "CR provavel / validar",
        "Qtd_CRs_historico",
        "CRs_onde_apareceu",
        "Total_Hrs_Trab",
        "Total_Hrs_Disp",
        "Total_Hrs_Manu",
        "Primeira_aparicao",
        "Ultima_aparicao",
        "Dias sem lancamento",
        "Status sugerido",
        "Cobra PDE?",
        "Status manual",
        "Observacao",
        "_eq_key",
    ]
    return master[ordered].sort_values("Equipamento").reset_index(drop=True), historico_cr


def prepare_master(df: pd.DataFrame, historico: pd.DataFrame | None = None, reference_date=None) -> pd.DataFrame:
    master = df.copy()
    master.columns = [normalize_column(col) for col in master.columns]

    rename_map = {}
    for col in master.columns:
        norm = normalize_text(col)
        if norm == "CR PROVAVEL / VALIDAR":
            rename_map[col] = "CR provavel / validar"
        elif norm == "QTD CRS NO HISTORICO":
            rename_map[col] = "Qtd_CRs_historico"
        elif norm == "CRS ONDE APARECEU":
            rename_map[col] = "CRs_onde_apareceu"
        elif norm == "TOTAL HRS TRAB":
            rename_map[col] = "Total_Hrs_Trab"
        elif norm == "TOTAL HRS DISP":
            rename_map[col] = "Total_Hrs_Disp"
        elif norm == "TOTAL HRS MANU":
            rename_map[col] = "Total_Hrs_Manu"
        elif norm == "ULTIMA APARICAO":
            rename_map[col] = "Ultima_aparicao"
        elif norm == "PRIMEIRA APARICAO":
            rename_map[col] = "Primeira_aparicao"
        elif norm == "DIAS SEM LANCAMENTO":
            rename_map[col] = "Dias sem lancamento"
        elif norm == "STATUS SUGERIDO":
            rename_map[col] = "Status sugerido"
        elif norm == "COBRA PDE?":
            rename_map[col] = "Cobra PDE?"
        elif norm == "STATUS MANUAL":
            rename_map[col] = "Status manual"
        elif norm == "OBSERVACAO":
            rename_map[col] = "Observacao"
    master = master.rename(columns=rename_map)

    if "Equipamento" not in master.columns:
        raise ValueError("A base mestre precisa ter a coluna Equipamento.")
    if "CR provavel / validar" not in master.columns and "CR" in master.columns:
        master["CR provavel / validar"] = master["CR"]

    master["Equipamento"] = master["Equipamento"].astype(str).str.strip()
    master["_eq_key"] = master["Equipamento"].map(normalize_text)

    defaults = {
        "CR provavel / validar": "",
        "Qtd_CRs_historico": 1,
        "CRs_onde_apareceu": "",
        "Total_Hrs_Trab": 0.0,
        "Total_Hrs_Disp": 0.0,
        "Total_Hrs_Manu": 0.0,
        "Status sugerido": "CANDIDATO ATIVO",
        "Cobra PDE?": "SIM",
        "Status manual": "",
        "Observacao": "",
        "Primeira_aparicao": pd.NaT,
        "Ultima_aparicao": pd.NaT,
        "Dias sem lancamento": pd.NA,
    }
    for col, default in defaults.items():
        if col not in master.columns:
            master[col] = default

    for col in ["Total_Hrs_Trab", "Total_Hrs_Disp", "Total_Hrs_Manu", "Qtd_CRs_historico"]:
        master[col] = master[col].map(coerce_number)
    for col in ["Primeira_aparicao", "Ultima_aparicao"]:
        master[col] = master[col].map(coerce_date)

    if historico is not None and master["Ultima_aparicao"].isna().all():
        generated, _ = build_master_base(historico, reference_date=reference_date)
        cols_to_merge = ["_eq_key", "Primeira_aparicao", "Ultima_aparicao", "Dias sem lancamento"]
        master = master.drop(columns=[c for c in cols_to_merge[1:] if c in master.columns], errors="ignore")
        master = master.merge(generated[cols_to_merge], on="_eq_key", how="left")

    reference = pd.to_datetime(reference_date, dayfirst=True, errors="coerce") if reference_date is not None else pd.NaT
    if pd.notna(reference):
        mask = master["Ultima_aparicao"].notna()
        master.loc[mask, "Dias sem lancamento"] = (
            reference.normalize() - pd.to_datetime(master.loc[mask, "Ultima_aparicao"]).dt.normalize()
        ).dt.days

    return master.drop_duplicates("_eq_key", keep="first").reset_index(drop=True)


def _is_active_status(row) -> bool:
    manual = normalize_text(row.get("Status manual", ""))
    cobra = normalize_text(row.get("Cobra PDE?", ""))
    if manual in {"INATIVO", "NAO", "NAO COBRAR", "N"}:
        return False
    if manual in {"ATIVO", "SIM", "COBRAR", "S"}:
        return True
    return cobra in {"SIM", "S", "COBRAR"}


def _active_status_mask(df: pd.DataFrame) -> pd.Series:
    manual = df.get("Status manual", pd.Series("", index=df.index)).map(normalize_text)
    cobra = df.get("Cobra PDE?", pd.Series("", index=df.index)).map(normalize_text)
    inactive = manual.isin({"INATIVO", "NAO", "NAO COBRAR", "N"})
    manual_active = manual.isin({"ATIVO", "SIM", "COBRAR", "S"})
    cobra_active = cobra.isin({"SIM", "S", "COBRAR"})
    return manual_active | (~inactive & cobra_active)


def compare_pde(master: pd.DataFrame, atual: pd.DataFrame, reference_date=None) -> PDEResult:
    master = prepare_master(master, reference_date=reference_date)
    atual = prepare_report(atual)
    if reference_date is None:
        reference_date = atual["Data(para)"].max()
    reference_date = pd.to_datetime(reference_date, dayfirst=True, errors="coerce")

    master["Ativo_para_cobranca"] = _active_status_mask(master)
    ativos = master[master["Ativo_para_cobranca"]].copy()

    atual_cr = (
        atual.groupby(["_eq_key", "_cr_key"], as_index=False)
        .agg(
            Equipamento=("Equipamento", _first_non_empty),
            CR_atual=("CR", _first_non_empty),
            Hrs_Trab_atual=("Hrs Trab", "sum"),
            Hrs_Disp_atual=("Hrs Disp", "sum"),
            Hrs_Manu_atual=("Hrs Manu", "sum"),
            Primeira_data_atual=("Data(de)", "min"),
            Ultima_data_atual=("Data(para)", "max"),
            Linhas_atual=("Equipamento", "count"),
        )
    )
    atual_eq = (
        atual_cr.groupby("_eq_key")
        .agg(
            Equipamento=("Equipamento", _first_non_empty),
            CRs_atuais=("CR_atual", lambda s: " | ".join(sorted(set(map(str, s))))),
            Qtd_CRs_atual=("CR_atual", "nunique"),
            Hrs_Trab_atual=("Hrs_Trab_atual", "sum"),
            Hrs_Disp_atual=("Hrs_Disp_atual", "sum"),
            Hrs_Manu_atual=("Hrs_Manu_atual", "sum"),
            Ultima_data_atual=("Ultima_data_atual", "max"),
        )
        .reset_index()
    )

    atual_keys = set(atual_eq["_eq_key"])
    ativos_lancaram = ativos[ativos["_eq_key"].isin(atual_keys)].merge(
        atual_eq[["_eq_key", "CRs_atuais", "Qtd_CRs_atual", "Ultima_data_atual"]],
        on="_eq_key",
        how="left",
    )
    cr_base = ativos_lancaram["CR provavel / validar"].map(normalize_text)
    crs_atuais = ativos_lancaram["CRs_atuais"].fillna("").map(
        lambda value: {normalize_text(item) for item in str(value).split("|")}
    )
    ativos_lancaram["Situacao"] = pd.Series(
        ["OK" if base in atuais else "LANCADO EM CR DIFERENTE" for base, atuais in zip(cr_base, crs_atuais)],
        index=ativos_lancaram.index,
    )
    ativos_lancaram = ativos_lancaram[
        ["Equipamento", "CR provavel / validar", "CRs_atuais", "Qtd_CRs_atual", "Ultima_data_atual", "Situacao"]
    ].rename(columns={"CR provavel / validar": "CR base/provavel"})

    pendencias = ativos[~ativos["_eq_key"].isin(atual_keys)].copy()
    pendencias["Acao recomendada"] = pendencias["Qtd_CRs_historico"].map(
        lambda qtd: "COBRAR" if int(qtd) == 1 else "VALIDAR ANTES DE COBRAR"
    )
    pendencias = pendencias[
        [
            "Equipamento",
            "CR provavel / validar",
            "Status sugerido",
            "Qtd_CRs_historico",
            "CRs_onde_apareceu",
            "Ultima_aparicao",
            "Dias sem lancamento",
            "Acao recomendada",
        ]
    ].rename(columns={"CR provavel / validar": "CR provavel"})

    pendencias_por_obra = (
        pendencias.groupby("CR provavel", dropna=False)
        .agg(
            Equipamentos_pendentes=("Equipamento", "count"),
            Maior_dias_sem_lancamento=("Dias sem lancamento", "max"),
            Menor_dias_sem_lancamento=("Dias sem lancamento", "min"),
        )
        .reset_index()
        .sort_values(["Equipamentos_pendentes", "CR provavel"], ascending=[False, True])
    )

    transferencias = ativos_lancaram[ativos_lancaram["Situacao"] == "LANCADO EM CR DIFERENTE"].copy()
    transferencias["Acao"] = "VALIDAR TRANSFERENCIA / ATUALIZAR BASE"

    validar = master[~master["Ativo_para_cobranca"]].copy()
    validar = validar.merge(
        atual_eq[["_eq_key", "CRs_atuais", "Qtd_CRs_atual"]],
        on="_eq_key",
        how="left",
    )
    has_cr_atual = validar["CRs_atuais"].fillna("").astype(str).str.strip().ne("")
    validar["Informacao extra"] = "VALIDAR STATUS / OBRA RESPONSAVEL"
    validar.loc[has_cr_atual, "Informacao extra"] = "USAR CR ATUAL PARA ATUALIZAR BASE"
    validar["Acao"] = ""
    validar_antes = validar[
        [
            "Equipamento",
            "CRs_onde_apareceu",
            "CRs_atuais",
            "Informacao extra",
            "Acao",
        ]
    ].rename(columns={"CRs_onde_apareceu": "CRs no historico", "CRs_atuais": "CR atual / ou totais"})

    base_all_keys = set(master["_eq_key"])
    ativos_keys = set(ativos["_eq_key"])
    fora = atual_eq[~atual_eq["_eq_key"].isin(ativos_keys)].copy()
    esta_na_base = fora["_eq_key"].isin(base_all_keys)
    fora["Esta na base?"] = np.where(esta_na_base, "SIM", "NAO")
    fora["Origem"] = np.where(esta_na_base, "Base em validacao", "Nao esta na base mestre")
    fora_base_ativa = fora[["Equipamento", "CRs_atuais", "Esta na base?", "Origem"]].rename(
        columns={"CRs_atuais": "CR atual no relatorio"}
    )

    duplicidades_atual = atual_eq[atual_eq["Qtd_CRs_atual"] > 1][
        ["Equipamento", "Qtd_CRs_atual", "CRs_atuais", "Hrs_Trab_atual", "Hrs_Disp_atual"]
    ].rename(columns={"Qtd_CRs_atual": "Qtd CRs atual", "CRs_atuais": "CRs atuais"})

    resumo_atual = (
        atual_cr.groupby("CR_atual")
        .agg(
            Equipamentos_unicos=("Equipamento", "nunique"),
            Hrs_Trab=("Hrs_Trab_atual", "sum"),
            Hrs_Disp=("Hrs_Disp_atual", "sum"),
            Hrs_Manu=("Hrs_Manu_atual", "sum"),
        )
        .reset_index()
        .rename(columns={"CR_atual": "CR / Obra"})
        .sort_values("Equipamentos_unicos", ascending=False)
    )

    dashboard_rows = [
        ("Data de referencia", reference_date.date() if pd.notna(reference_date) else ""),
        ("Equipamentos na base mestre", len(master)),
        ("Equipamentos ativos para cobranca", len(ativos)),
        ("Equipamentos unicos no relatorio atual", atual_eq["_eq_key"].nunique()),
        ("Ativos que lancaram no periodo", len(ativos_lancaram)),
        ("Ativos pendentes", len(pendencias)),
        ("Obras com pendencia", pendencias["CR provavel"].nunique() if not pendencias.empty else 0),
        ("Equipamentos para validar antes de cobrar", len(validar_antes)),
        ("Possiveis transferencias", len(transferencias)),
        ("Equipamentos fora da base ativa", len(fora_base_ativa)),
        ("Duplicidades no atual", len(duplicidades_atual)),
    ]
    dashboard = pd.DataFrame(dashboard_rows, columns=["Indicador", "Valor"])

    return PDEResult(
        dashboard=dashboard,
        pendencias=pendencias.sort_values(["CR provavel", "Equipamento"]).reset_index(drop=True),
        pendencias_por_obra=pendencias_por_obra.reset_index(drop=True),
        ativos_lancaram=ativos_lancaram.sort_values("Equipamento").reset_index(drop=True),
        validar_antes=validar_antes.sort_values("Equipamento").reset_index(drop=True),
        transferencias=transferencias.reset_index(drop=True),
        fora_base_ativa=fora_base_ativa.sort_values("Equipamento").reset_index(drop=True),
        duplicidades_atual=duplicidades_atual.sort_values("Equipamento").reset_index(drop=True),
        resumo_atual=resumo_atual.reset_index(drop=True),
        base_mestre=master.drop(columns=["Ativo_para_cobranca"], errors="ignore").sort_values("Equipamento").reset_index(drop=True),
        historico_cr=pd.DataFrame(),
    )


def build_pde_report(
    historico_paths: Iterable[str | Path],
    atual_path: str | Path,
    base_mestre_path: str | Path | None = None,
    reference_date=None,
) -> PDEResult:
    historico = read_reports(historico_paths)
    atual = read_table(atual_path)
    atual_prepared = prepare_report(atual)
    if reference_date is None:
        reference_date = atual_prepared["Data(para)"].max()

    generated_master, historico_cr = build_master_base(historico, reference_date=reference_date)
    if base_mestre_path:
        master = prepare_master(find_master_sheet(base_mestre_path), historico=historico, reference_date=reference_date)
    else:
        master = generated_master

    result = compare_pde(master, atual_prepared, reference_date=reference_date)
    return PDEResult(
        dashboard=result.dashboard,
        pendencias=result.pendencias,
        pendencias_por_obra=result.pendencias_por_obra,
        ativos_lancaram=result.ativos_lancaram,
        validar_antes=result.validar_antes,
        transferencias=result.transferencias,
        fora_base_ativa=result.fora_base_ativa,
        duplicidades_atual=result.duplicidades_atual,
        resumo_atual=result.resumo_atual,
        base_mestre=result.base_mestre,
        historico_cr=historico_cr.drop(columns=["_eq_key", "_cr_key"], errors="ignore"),
    )


def _safe_sheet_name(name: str, used: set[str]) -> str:
    cleaned = str(name)
    for ch in r"\/?*[]:":
        cleaned = cleaned.replace(ch, " ")
    cleaned = re.sub(r"\s+", " ", cleaned).strip() or "Aba"
    cleaned = cleaned[:31]
    candidate = cleaned
    index = 2
    while candidate in used:
        suffix = f" {index}"
        candidate = f"{cleaned[:31 - len(suffix)]}{suffix}"
        index += 1
    used.add(candidate)
    return candidate


def _excel_value(value):
    if pd.isna(value):
        return None
    if isinstance(value, pd.Timestamp):
        return value.to_pydatetime()
    if isinstance(value, (datetime, date)):
        return value
    return value


def _write_df(ws, df: pd.DataFrame, title: str, table_name: str, note: str | None = None):
    palette = {
        "title": "1E2A45",
        "header": "29415F",
        "header_font": "FFFFFF",
        "accent": "E8F1F8",
        "border": "D9E2EC",
        "danger": "FCE8E6",
        "warn": "FFF4D6",
        "ok": "E6F4EA",
        "total": "DDE7F0",
    }
    thin = Side(style="thin", color=palette["border"])
    border = Border(bottom=thin)

    cols = list(df.columns) if not df.empty else list(df.columns)
    ncols = max(len(cols), 1)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    title_cell = ws.cell(1, 1, title)
    title_cell.font = Font("Segoe UI", bold=True, size=14, color="FFFFFF")
    title_cell.fill = PatternFill("solid", fgColor=palette["title"])
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 30

    row_cursor = 2
    if note:
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
        note_cell = ws.cell(2, 1, note)
        note_cell.font = Font("Segoe UI", italic=True, size=10, color="475569")
        note_cell.fill = PatternFill("solid", fgColor=palette["accent"])
        row_cursor = 3

    header_row = row_cursor
    for col_idx, col_name in enumerate(cols, 1):
        cell = ws.cell(header_row, col_idx, col_name)
        cell.font = Font("Segoe UI", bold=True, color=palette["header_font"])
        cell.fill = PatternFill("solid", fgColor=palette["header"])
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    ws.row_dimensions[header_row].height = 24

    col_meta = {
        col_name: {
            "is_text": df[col_name].dtype == "object",
            "wrap": col_name in {"CRs onde apareceu", "CRs no historico", "CR atual / ou totais"},
            "norm": normalize_text(col_name),
        }
        for col_name in cols
    }

    for row_idx, row in enumerate(df.itertuples(index=False, name=None), header_row + 1):
        for col_idx, (col_name, raw_value) in enumerate(zip(cols, row), 1):
            value = _excel_value(raw_value)
            meta = col_meta[col_name]
            cell = ws.cell(row_idx, col_idx, value)
            cell.font = Font("Segoe UI", size=10)
            cell.border = border
            cell.alignment = Alignment(
                horizontal="left" if meta["is_text"] else "right",
                vertical="center",
                wrap_text=meta["wrap"],
            )
            col_norm = meta["norm"]
            if "DATA" in col_norm or "APARICAO" in col_norm:
                cell.number_format = "yyyy-mm-dd"
            elif "HRS" in col_norm or "TOTAL" in col_norm or "DIAS" in col_norm or "QTD" in col_norm:
                cell.number_format = "#,##0.00" if "HRS" in col_norm else "#,##0"

            if "ACAO" in col_norm or "STATUS" in col_norm or "SITUACAO" in col_norm:
                value_norm = normalize_text(value)
                if any(term in value_norm for term in ["COBRAR", "PENDENTE", "DIFERENTE"]):
                    cell.fill = PatternFill("solid", fgColor=palette["danger"])
                elif "VALIDAR" in value_norm:
                    cell.fill = PatternFill("solid", fgColor=palette["warn"])
                elif value_norm == "OK":
                    cell.fill = PatternFill("solid", fgColor=palette["ok"])

    last_row = max(header_row + len(df), header_row)
    if cols and len(df) > 0:
        end_col = get_column_letter(len(cols))
        ref = f"A{header_row}:{end_col}{last_row}"
        table = Table(displayName=table_name, ref=ref)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        ws.add_table(table)

    for col_idx, col_name in enumerate(cols, 1):
        values = [str(col_name)]
        if not df.empty:
            values.extend(str(v) for v in df[col_name].head(200).fillna(""))
        width = min(max(max(len(v) for v in values) + 3, 12), 48)
        if col_name in {"CRs onde apareceu", "CRs no historico", "CR atual / ou totais"}:
            width = 55
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = ws.cell(header_row + 1, 1).coordinate
    ws.sheet_view.showGridLines = False
    return ws


def _build_dashboard(ws, result: PDEResult):
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:F1")
    ws["A1"] = "RELATORIO DE PENDENCIAS DE PDE"
    ws["A1"].font = Font("Segoe UI", bold=True, size=16, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="1E2A45")
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 34

    ws["A3"] = "Indicador"
    ws["B3"] = "Valor"
    for cell in ws["3:3"]:
        cell.font = Font("Segoe UI", bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="29415F")
        cell.alignment = Alignment(horizontal="center")

    for row_idx, (_, row) in enumerate(result.dashboard.iterrows(), 4):
        ws.cell(row_idx, 1, row["Indicador"])
        ws.cell(row_idx, 2, _excel_value(row["Valor"]))
        ws.cell(row_idx, 1).font = Font("Segoe UI", size=10)
        ws.cell(row_idx, 2).font = Font("Segoe UI", bold=True, size=10)
        ws.cell(row_idx, 1).alignment = Alignment(horizontal="left")
        ws.cell(row_idx, 2).alignment = Alignment(horizontal="right")
        if row["Indicador"] in {"Ativos pendentes", "Possiveis transferencias", "Duplicidades no atual"}:
            ws.cell(row_idx, 2).fill = PatternFill("solid", fgColor="FCE8E6")

    start = 4 + len(result.dashboard) + 3
    ws.cell(start, 1, "Pendencias por obra")
    ws.cell(start, 1).font = Font("Segoe UI", bold=True, size=12, color="1E2A45")
    top = result.pendencias_por_obra.head(10)
    headers = ["CR provavel", "Equipamentos pendentes", "Maior dias sem lancamento"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(start + 1, col_idx, header)
        cell.font = Font("Segoe UI", bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="29415F")
    for offset, (_, row) in enumerate(top.iterrows(), start + 2):
        ws.cell(offset, 1, row.get("CR provavel"))
        ws.cell(offset, 2, row.get("Equipamentos_pendentes"))
        ws.cell(offset, 3, row.get("Maior_dias_sem_lancamento"))

    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 24
    ws.freeze_panes = "A4"


def export_pde_excel(result: PDEResult, output_path: str | Path) -> Path:
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    used: set[str] = set()
    ws = wb.active
    ws.title = _safe_sheet_name("Dashboard", used)
    _build_dashboard(ws, result)

    sheet_specs = [
        ("Pendencias por Obra", result.pendencias_por_obra, "PendenciasPorObra", "Resumo de cobranca por obra/CR."),
        ("Pendencias para Cobranca", result.pendencias, "PendenciasCobranca", "Equipamentos ativos esperados que nao apareceram no relatorio atual."),
        ("Ativos que Lancaram", result.ativos_lancaram, "AtivosLancaram", "Equipamentos da base ativa encontrados no periodo atual."),
        ("Validar antes de Cobrar", result.validar_antes, "ValidarAntesCobrar", "Equipamentos com historico em multiplos CRs ou status de validacao."),
        ("Possiveis Transferencias", result.transferencias, "PossiveisTransferencias", "Ativos encontrados no relatorio atual em CR diferente da base."),
        ("Lancaram Fora da Base Ativa", result.fora_base_ativa, "ForaBaseAtiva", "Equipamentos no atual que nao fazem parte da base ativa de cobranca."),
        ("Duplicidades Atual", result.duplicidades_atual, "DuplicidadesAtual", "Equipamentos que apareceram em mais de um CR no relatorio atual."),
        ("Resumo Relatorio Atual", result.resumo_atual, "ResumoAtual", "Totais por obra/CR do relatorio atual."),
        ("Base Mestre Usada", result.base_mestre.drop(columns=["_eq_key"], errors="ignore"), "BaseMestreUsada", "Base considerada para a comparacao."),
    ]
    if not result.historico_cr.empty:
        sheet_specs.append(
            ("Equipamento x CR Historico", result.historico_cr, "HistoricoCR", "Consolidacao historica por equipamento e CR.")
        )

    for name, df, table_name, note in sheet_specs:
        sheet = wb.create_sheet(_safe_sheet_name(name, used))
        _write_df(sheet, df, name.upper(), table_name, note)

    wb.save(output_path)
    return output_path
